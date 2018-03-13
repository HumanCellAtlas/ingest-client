#!/usr/bin/env python
"""
This script will read a spreadsheet, generate a manifest, submit all items to the ingest API, 
assign uuid and generate a directory of bundles for the submitted data
"""
from openpyxl.utils.exceptions import InvalidFileException

from spreadsheetUploadError import SpreadsheetUploadError

__author__ = "jupp"
__license__ = "Apache 2.0"

import json, os
from openpyxl import load_workbook
from ingestapi import IngestApi
from optparse import OptionParser
import logging

from itertools import chain
from collections import defaultdict

# these are spreadsheet fields that can be a list
# todo - these should be read out of the json schema at the start
# schema_ontologyFields = {"donor" : ["ancestry", "development_stage", "disease", "medication", "strain", "genus_species"],
#                     "cell_suspension" : ["target_cell_type", "genus_species"],
#                     "death" : ["cause_of_death"],
#                     "immortalized_cell_line" : ["cell_type", "disease", "cell_cycle", "genus_species"],
#                     "protocol" : ["type"],
#                     "primary_cell_line" : ["cell_type", "disease", "cell_cycle", "genus_species"],
#                     "specimen_from_organism" : ["body_part", "organ", "genus_species"],
#                     "project" : ["experimental_design"],
#                     "organoid" : ["model_for_organ", "genus_species"]
#                     }

schema_arrayFields = {
    "cell_line": ["genus_species", "publications", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "cell_suspension": ["genus_species", "target_cell_type", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "donor_organism": ["genus_species", "disease", "familial_relationship", "ancestry", "strain", "ncbi_taxon_id",
                       "supplementary_files", "process_ids"],
    "organoid": ["genus_species", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "specimen_from_organism": ["genus_species", "disease", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "sequence_file": ["insdc_run"],
    "analysis_process": ["inputs", "tasks", "input_bundles", "outputs", "operator_identity"],
    "collection_process": ["process_reagents", "operator_identity"],
    "dissociation_process": ["process_reagents", "operator_identity"],
    "imaging_process": ["field_counts", "field_microns", "field_resolution", "operator_identity"],
    "project": ["contributors", "supplementary_files", "publications"],
    "publication": ["authors"]
}

schema_timeFields = {
    "cell_line": ["date_established"],
    "donor_organism": ["time_of_death"],
    "specimen_from_organism": ["collection_time"],
    "analysis_process": ["start_time", "stop_time", "timestamp_start_utc", "timestamp_stop_utc"],
    "collection_process": ["start_time"],
    "dissociation_process": ["start_time"],
    "enrichment_process": ["start_time"],
    "imaging_process": ["start_time"],
    "library_preparation_process": ["start_time"],
    "sequencing_process": ["start_time"]
}

schema_booleanFields = {
    "donor_organism": ["is_living"],
    "sequencing_process": ["paired_ends"],
    "death": ["cold_perfused"]
}

schema_integerFields = {
    "cell_line": ["ncbi_taxon_id", "passage_number"],
    "cell_suspension": ["total_estimated_cells", "ncbi_taxon_id", "passage_number"],
    "donor_organism": ["ncbi_taxon_id", "hardy_scale"],
    "organoid": ["ncbi_taxon_id", "passage_number"],
    "specimen_from_organism": ["ncbi_taxon_id", "ischemic_time"],
    "sequence_file": ["lane_index", "read_length"],
    "analysis_process": ["cpus"],
    "imaging_process": ["field_counts", "field_microns", "field_resolution"],
    "library_preparation_process": ["spike_in_dilution", "barcode_offset", "barcode_length"],
    "project": ["pmid"]
}

# maps sheets to the latest version of each schema
# todo - this should be replaced by dynamic lookups against ingest-core /schemas endpoint
# (spec for lookup: query by 'title' (sheet name in the spreadsheet, field in the schema)
schema_sheetname_mappings = {
    # biomaterials
    "cell_line": "https://schema.humancellatlas.org/type/biomaterial/5.0.1/cell_line",
    "cell_suspension": "https://schema.humancellatlas.org/type/biomaterial/5.0.0/cell_suspension",
    "donor_organism": "https://schema.humancellatlas.org/type/biomaterial/5.0.0/donor_organism",
    "organoid": "https://schema.humancellatlas.org/type/biomaterial/5.0.0/organoid",
    "specimen_from_organism": "https://schema.humancellatlas.org/type/biomaterial/5.0.0/specimen_from_organism",
    # files
    "analysis_file": "https://schema.humancellatlas.org/type/file/5.0.0/analysis_file",
    "sequence_file": "https://schema.humancellatlas.org/type/file/5.0.0/sequence_file",
    # analysis processes
    "analysis_process": "https://schema.humancellatlas.org/type/process/analysis/5.0.0/analysis_process",
    # biomaterial processes
    "collection_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.0.0/collection_process",
    "dissociation_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.0.0/dissociation_process",
    "enrichment_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.0.0/enrichment_process",
    # imaging processes
    "imaging_process": "https://schema.humancellatlas.org/type/process/imaging/5.0.0/imaging_process",
    # sequencing processes
    "library_preparation_process": "https://schema.humancellatlas.org/type/process/sequencing/5.0.0/library_preparation_process",
    "sequencing_process": "https://schema.humancellatlas.org/type/process/sequencing/5.0.0/sequencing_process",
    # projects
    "project": "https://schema.humancellatlas.org/type/project/5.0.1/project",
    # protocols
    "protocol": "https://schema.humancellatlas.org/type/protocol/5.0.0/protocol",
    "analysis_protocol": "https://schema.humancellatlas.org/type/protocol/analysis/5.0.0/analysis_protocol",
    "biomaterial_collection_protocol": "https://schema.humancellatlas.org/type/protocol/biomaterial/5.0.0/biomaterial_collection_protocol",
    "imaging_protocol": "https://schema.humancellatlas.org/type/protocol/imaging/5.0.0/imaging_protocol",
    "sequencing_protocol": "https://schema.humancellatlas.org/type/protocol/sequencing/5.0.0/sequencing_protocol"
}

# schema_fieldname_mappings = {
#     # cell_line
#     "biomaterial_core": "https://schema.humancellatlas.org/core/biomaterial/5.0.0/biomaterial_core",
#     "cell_cycle": "https://schema.humancellatlas.org/module/ontology/5.0.0/cell_cycle_ontology",
#     "cell_morphology": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/cell_morphology",
#     "growth_conditions": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/growth_conditions",
#     "cell_type": "https://schema.humancellatlas.org/module/ontology/5.0.0/cell_type_ontology",
#     "disease": "https://schema.humancellatlas.org/module/ontology/5.0.0/disease_ontology",
#     "genus_species": "https://schema.humancellatlas.org/module/ontology/5.0.0/species_ontology",
#     "publications": "https://schema.humancellatlas.org/module/project/5.0.1/publication",
#     # cell_suspension
#     "target_cell_type": "https://schema.humancellatlas.org/module/ontology/5.0.0/cell_type_ontology",
#     # donor_organism
#     "human_specific": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/homo_sapiens_specific",
#     "mus_musculus_specific": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/mus_musculus_specific",
#     "death": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/death",
#     "medical_history": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/medical_history",
#     "organism_age_unit": "https://schema.humancellatlas.org/module/ontology/5.0.0/time_unit_ontology",
#     "development_stage": "https://schema.humancellatlas.org/module/ontology/5.0.0/development_stage_ontology",
#     "familial_relationship": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/familial_relationship",
#     "gestational_age_unit": "https://schema.humancellatlas.org/module/ontology/5.0.0/time_unit_ontology",
#     "height_unit": "https://schema.humancellatlas.org/module/ontology/5.0.0/length_unit_ontology",
#     "weight_unit": "https://schema.humancellatlas.org/module/ontology/5.0.0/mass_unit_ontology",
#     # organoid
#     "model_for_organ": "https://schema.humancellatlas.org/module/ontology/5.0.0/organ_ontology",
#     "organoid_age_unit": "https://schema.humancellatlas.org/module/ontology/5.0.0/time_unit_ontology",
#     # specimen_from_organism
#     "organ": "https://schema.humancellatlas.org/module/ontology/5.0.0/organ_ontology",
#     "organ_part": "https://schema.humancellatlas.org/module/ontology/5.0.0/organ_part_ontology",
#     "state_of_specimen": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/state_of_specimen",
#     "preservation_storage": "https://schema.humancellatlas.org/module/biomaterial/5.0.0/preservation_storage",
#     # file
#     "file_core": "https://schema.humancellatlas.org/core/file/5.0.0/file_core",
#     # process
#     "process_core": "https://schema.humancellatlas.org/core/process/5.0.0/process_core",
#     "process_type": "https://schema.humancellatlas.org/module/ontology/5.0.0/process_type_ontology",
#     "outputs.items": "https://schema.humancellatlas.org/type/file/5.0.0/analysis_file",
#     # collection_process
#     "process_reagents": "https://schema.humancellatlas.org/module/process/5.0.0/purchased_reagents",
#     # library_preparation_process
#     "cell_barcode": "https://schema.humancellatlas.org/module/process/sequencing/5.0.0/barcode",
#     "input_nucleic_acid_molecule": "https://schema.humancellatlas.org/module/ontology/5.0.0/biological_macromolecule_ontology",
#     # sequencing_process
#     "instrument_manufacturer_model": "https://schema.humancellatlas.org/module/ontology/5.0.0/instrument_ontology",
#     "smartseq2": "https://schema.humancellatlas.org/module/process/sequencing/5.0.0/smartseq2",
#     # project
#     "project_core": "https://schema.humancellatlas.org/core/project/5.0.0/project_core",
#     "contributors": "https://schema.humancellatlas.org/module/project/5.0.0/contact",
#     # analysis_protocol
#     "protocol_core": "https://schema.humancellatlas.org/core/protocol/5.0.0/protocol_core",
#     "protocol_type": "https://schema.humancellatlas.org/module/ontology/5.0.0/protocol_type_ontology"
# }

# SCHEMA_URL = os.environ.get('SCHEMA_URL', "https://raw.githubusercontent.com/HumanCellAtlas/metadata-schema/%s/json_schema/")
# SCHEMA_URL = os.path.expandvars(os.environ.get('SCHEMA_URL', SCHEMA_URL))
# SCHEMA_VERSION = os.environ.get('SCHEMA_VERSION', '4.6.1')


class SpreadsheetSubmission:

    def __init__(self, dry=False, output=None, schema_version=None):
        # formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        # logging.basicConfig(formatter=formatter, level=logging.INFO)
        self.logger = logging.getLogger(__name__)

        self.dryrun = dry
        self.outputDir = output
        self.ingest_api = None
        # self.schema_version = schema_version if schema_version else os.path.expandvars(SCHEMA_VERSION)
        # self.schema_url = os.path.expandvars(SCHEMA_URL % self.schema_version)
        if not self.dryrun:
            self.ingest_api = IngestApi()

    def createSubmission(self, token):
        self.logger.info("creating submission...")
        if not self.ingest_api:
            self.ingest_api = IngestApi()

        submissionUrl = self.ingest_api.createSubmission(token)
        self.logger.info("new submission " + submissionUrl)
        return submissionUrl

    # this only works for keys nested to two levels, need something smarter to handle arbitrary
    # depth keys e.g. we support <level1>.<level2> = <value>, where value is either a single value or list
    # this doesn't support <level1>.<level2>.<level3> except if <level3> is "ontology"
    #This function takes a dictionary object, a key/value pair from the spreadsheet input and an object type, and adds
    #the key/value pair to the dictionary. The type is used to assess whether a field is an ontology or an array based on the
    # two lookup dictionaries declared at the top.
    # For || separated strings, the value field is split. For json arrays, the value is put into an array. Json objects of type
    # ontology require some extra formatting, which is also done here.

    def _keyValueToNestedObject(self, obj, key, value, type):
        if "*" in key:
            key = key.replace("*", "")
        d = str(value)
        # If the value contains a double pipe (||) or the key is for a field that can be a list (with or without also being
        # an ontology field), put value into an array (splitting if necessary)
        if "\"" in str(value) or "||" in str(value) \
                or (type in schema_arrayFields.keys() and key.split('.')[-1] in schema_arrayFields[type]) \
                or (type in schema_arrayFields.keys() and key.split('.')[-1] == "ontology"
                    and key.split('.')[-2] in schema_arrayFields[type]):
            if "||" in str(value):
            # d = map(lambda it: it.strip(' "\''), str(value).split("||"))
                d = str(value).split("||")
            else:
                d = [value]

        # Raise an error if the key is too nested
        if len(key.split('.')) > 3:
            raise ValueError('We don\'t support keys nested greater than 3 levels, found:'+key)

        # If the key is in the date_time field list, convert the date time into a string of format YYYY-MM-DDThh:mm:ssZ
        # so it validates
        if type in schema_timeFields.keys():
            try:
                if key.split('.')[-1] in schema_timeFields[type]:
                    if isinstance(d, list):
                        for i, v in enumerate(d):
                            date_string = v.strftime("%Y-%m-%dT%H:%M:%SZ")
                            d[i] = date_string
                    else:
                        d = d.strftime("%Y-%m-%dT%H:%M:%SZ")
            except:
                self.logger.warn('Failed to convert field %s (value %s) to date_time value'.format(type, d))
        elif type in schema_integerFields.keys():
            if key.split('.')[-1] in schema_integerFields[type]:
                try:
                    d = int(d)
                except:
                    self.logger.warn('Failed to convert field %s (value %s) to integer value'.format(type, d))
                    d = str(d)
        elif type in schema_booleanFields.keys():
            if key.split('.')[-1] in schema_booleanFields[type]:
                try:
                    if d.lower() in ["true", "yes"]:
                        d = True
                    elif d in ["false", "no"]:
                        d = False
                except:
                    self.logger.warn('Failed to convert field %s (value %s) to integer value'.format(type, d))
                    d = str(d)
        else:
            d = str(d)

        # If the key is in the ontology field list, or contains "ontology", format it according to the ontology json schema
        # if type in schema_ontologyFields.keys():
        #     if key.split('.')[-1] in schema_ontologyFields[type]:
        #         if isinstance(d, list):
        #             t = []
        #             for index, v in enumerate(d):
        #                 t.append({"text" : d[index]})
        #             d = t
        #         else:
        #             d = {"text" : d}
        #     elif key.split('.')[-1] == "ontology":
        #         if isinstance(d, list):
        #             t = []
        #             for index, v in enumerate(d):
        #                 t.append({"ontology": d[index]})
        #             d = t
        #         else:
        #             d = {"ontology": d}
        #         key = ".". join(key.split('.')[:-1])



        # Build up the key-value dictionary
        for part in reversed(key.split('.')):
            d = {part: d}

        return self._mergeDict(obj, d)

    def _mergeDict(self, dict1, dict2):
        dict3 = defaultdict(list)
        for k, v in chain(dict1.items(), dict2.items()):
            if k in dict3:
                if isinstance(v, dict):
                    dict3[k].update(self._mergeDict(dict3[k], v))
                elif isinstance(v, list) and isinstance(dict3[k], list) and len(v) == len(dict3[k]):
                    for index, e in enumerate(v):
                        dict3[k][index].update(self._mergeDict(dict3[k][index], e))
                else:
                    dict3[k].update(v)
            else:
                dict3[k] = v
        return dict3

    #sheets with one or more data rows and properties in row 1
    def _multiRowToObjectFromSheet(self, type, sheet):
        objs = []
        for row in sheet.iter_rows(row_offset=3, max_row=(sheet.max_row - 3)):
            obj = {}
            hasData = False
            for cell in row:
                # WARNING: remove long from the checks if using python 3!
                if not cell.value and not isinstance(cell.value, (int, float, long)):
                    continue
                hasData = True
                cellCol = cell.col_idx
                propertyValue = sheet.cell(row=3, column=cellCol).value

                d = self._keyValueToNestedObject(obj, propertyValue, cell.value, type)
                obj.update(d)
            if hasData:
                self.logger.debug(json.dumps(obj))
                objs.append(obj)

        return objs

    # sheets that represent one entity where the properties are in column 0
    def _sheetToObject(self, type, sheet):
        obj = {}
        for row in sheet.iter_rows():
            if len(row) > 1:
                propertyCell = row[0].value
                valueCell = row[1].value
                if valueCell:
                    obj = self._keyValueToNestedObject(obj, propertyCell, valueCell, type)
                    # obj.update(d)
        self.logger.debug(json.dumps(obj))
        return obj

    def completeSubmission(self):
        self.ingest_api.finishSubmission()

    def submit(self, pathToSpreadsheet, submissionUrl, token=None, project_id=None):
        try:
            self._process(pathToSpreadsheet, submissionUrl, token, project_id)
        except ValueError as e:
            self.logger.error("Error:"+str(e))
            raise e

    def dumpJsonToFile(self, object, projectId, name):
        if self.outputDir:
            dir = os.path.abspath(self.outputDir)
            if not os.path.exists(dir):
                os.makedirs(dir)
            tmpFile = open(dir + "/" + str(projectId) + "_" + str(name) + ".json", "w")
            tmpFile.write(json.dumps(object, indent=4))
            tmpFile.close()

    def _process(self, pathToSpreadsheet, submissionUrl, token, existing_project_id):

        # parse the spreadsheet
        try:
            wb = load_workbook(filename=pathToSpreadsheet)
        except InvalidFileException:
            raise SpreadsheetUploadError(400, "The uploaded file is not a valid XLSX spreadsheet", "")
        # This code section deals with cases where the project has already been submitted
        # ASSUMPTION: now additional project information (publications, contributors etc) is added via
        # the spreadsheet if the project already exists

        projectSheet = wb.create_sheet()
        projectPubsSheet = wb.create_sheet()
        contributorSheet = wb.create_sheet()

        if existing_project_id is None:
            projectSheet = wb.get_sheet_by_name("project")

            if "project.publications" in wb.sheetnames:
                projectPubsSheet = wb.get_sheet_by_name("project.publications")
            if "contact" in wb.sheetnames:
                contributorSheet = wb.get_sheet_by_name("contact")

        specimenSheet = wb.create_sheet()
        donorSheet = wb.create_sheet()
        cellSuspensionSheet = wb.create_sheet()
        familialRelationshipSheet = wb.create_sheet()

        if "specimen_from_organism" in wb.sheetnames:
            specimenSheet = wb.get_sheet_by_name("specimen_from_organism")
        if "donor_organism" in wb.sheetnames:
            donorSheet = wb.get_sheet_by_name("donor_organism")
        if "cell_suspension" in wb.sheetnames:
            cellSuspensionSheet = wb.get_sheet_by_name("cell_suspension")
        if "familial_relationship" in wb.sheetnames:
            familialRelationshipSheet = wb.get_sheet_by_name("familial_relationship")

        organoidSheet = wb.create_sheet()
        if "organoid" in wb.sheetnames:
            organoidSheet = wb.get_sheet_by_name("organoid")

        clSheet = wb.create_sheet()
        if "cell_line" in wb.sheetnames:
            clSheet = wb.get_sheet_by_name("cell_line")

        clPublicationSheet = wb.create_sheet()
        if "cell_line.publications" in wb.sheetnames:
            clPublicationSheet = wb.get_sheet_by_name("cell_line.publications")


        protocolSheet = wb.create_sheet()
        collectionSheet = wb.create_sheet()
        dissociationSheet = wb.create_sheet()
        enrichmentSheet = wb.create_sheet()
        libraryPrepSheet = wb.create_sheet()
        sequencingSheet = wb.create_sheet()
        reagentsSheet = wb.create_sheet()
        filesSheet = wb.create_sheet()

        if "protocol" in wb.sheetnames:
            protocolSheet = wb.get_sheet_by_name("protocol")
        if "enrichment_process" in wb.sheetnames:
            enrichmentSheet = wb.get_sheet_by_name("enrichment_process")
        if "collection_process" in wb.sheetnames:
            collectionSheet = wb.get_sheet_by_name("collection_process")
        if "dissociation_process" in wb.sheetnames:
            dissociationSheet = wb.get_sheet_by_name("dissociation_process")
        if "library_preparation_process" in wb.sheetnames:
            libraryPrepSheet = wb.get_sheet_by_name("library_preparation_process")
        if "sequencing_process" in wb.sheetnames:
            sequencingSheet = wb.get_sheet_by_name("sequencing_process")
        if "purchased_reagents" in wb.sheetnames:
            reagentsSheet = wb.get_sheet_by_name("purchased_reagents")
        if "sequence_file" in wb.sheetnames:
            filesSheet = wb.get_sheet_by_name("sequence_file")


        # convert data in sheets back into dict
        project = self._multiRowToObjectFromSheet("project", projectSheet)
        if project:
            if len(project) == 1:
                project = project[0]
                project.update({"type": "project",
                  "describedBy": schema_sheetname_mappings["project"]})

        enrichment = self._multiRowToObjectFromSheet("enrichment_process", enrichmentSheet)
        if enrichment:
            for e in enrichment:
                e.update({"type": "process",
                          "describedBy": schema_sheetname_mappings["enrichment_process"]})

        collection = self._multiRowToObjectFromSheet("collection_process", collectionSheet)
        if collection:
            for c in collection:
                c.update({"type": "process",
                     "describedBy": schema_sheetname_mappings["collection_process"]})

        dissociation = self._multiRowToObjectFromSheet("dissociation_process", dissociationSheet)
        if dissociation:
            for d in dissociation:
                d.update({"type": "process",
                  "describedBy": schema_sheetname_mappings["dissociation_process"]})

        reagents = self._multiRowToObjectFromSheet("purchased_reagents", reagentsSheet)
        libraryPrep = self._multiRowToObjectFromSheet("library_preparation_process", libraryPrepSheet)
        if libraryPrep:
            for lp in libraryPrep:
                lp.update({"type": "process",
                    "describedBy": schema_sheetname_mappings["library_preparation_process"]})
        sequencing = self._multiRowToObjectFromSheet("sequencing_process", sequencingSheet)
        if sequencing:
            for s in sequencing:
                s.update({"type": "process",
                  "describedBy": schema_sheetname_mappings["sequencing_process"]})

        protocols = self._multiRowToObjectFromSheet("protocol", protocolSheet)
        if protocols:
            for prot in protocols:
                prot.update({"type": "protocol",
                  "describedBy": schema_sheetname_mappings["protocol"]})
        donors = self._multiRowToObjectFromSheet("donor_organism", donorSheet)
        if donors:
            for do in donors:
                do.update({"type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["donor_organism"]})
        familialRelationships = self._multiRowToObjectFromSheet("familial_relationship", familialRelationshipSheet)
        publications = self._multiRowToObjectFromSheet("project.publications", projectPubsSheet)
        contributors = self._multiRowToObjectFromSheet("contributor", contributorSheet)

        specimens = self._multiRowToObjectFromSheet("specimen_from_organism", specimenSheet)
        if specimens:
            for spec in specimens:
                spec.update({"type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["specimen_from_organism"]})
        cell_suspension = self._multiRowToObjectFromSheet("cell_suspension", cellSuspensionSheet)
        if cell_suspension:
            for cs in cell_suspension:
                cs.update({"type": "biomaterial",
                     "describedBy": schema_sheetname_mappings["cell_suspension"]})
        organoid = self._multiRowToObjectFromSheet("organoid", organoidSheet)
        if organoid:
            for org in organoid:
                org.update({"type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["organoid"]})
        cell_line = self._multiRowToObjectFromSheet("cell_line", clSheet)
        if cell_line:
            for cl in cell_line:
                cl.update({"type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["cell_line"]})
        cell_line_publications = self._multiRowToObjectFromSheet("cell_line.publications", clPublicationSheet)
        files = self._multiRowToObjectFromSheet("sequence_file", filesSheet)
        if files:
            for f in files:
                f.update({"type": "file",
                 "describedBy": schema_sheetname_mappings["sequence_file"]})


        biomaterials = []
        biomaterials.extend(donors)
        biomaterials.extend(specimens)
        biomaterials.extend(cell_suspension)
        biomaterials.extend(organoid)
        biomaterials.extend(cell_line)

        processes = []
        processes.extend(collection)
        processes.extend(dissociation)
        processes.extend(enrichment)
        processes.extend(libraryPrep)
        processes.extend(sequencing)

        # creating submission
        #
        if not self.dryrun and not submissionUrl:
            submissionUrl = self.createSubmission(token)

        linksList = []

        # post objects to the Ingest API after some basic validation
        if existing_project_id is None:
            self.logger.info("Creating a new project for the submission")
            if "project_shortname" not in project["project_core"]:
                raise ValueError('Project must have an id attribute')
            projectId = project["project_core"]["project_shortname"]

             # embedd contact & publication into into project for now
            pubs = []
            for index, publication in enumerate(publications):
                pubs.append(publication)
            project["publications"] = pubs

            cont = []
            for index, contributor in enumerate(contributors):
                cont.append(contributor)
            project["contributors"] = cont

            self.dumpJsonToFile(project, projectId, "project")

            projectIngest = None
            if not self.dryrun:
                projectIngest = self.ingest_api.createProject(submissionUrl, json.dumps(project), token)

        else:
            if not self.dryrun:
                self.logger.info("Retreiving existing project: " + existing_project_id)
                projectIngest = self.ingest_api.getProjectById(existing_project_id)
                submissionEnvelope = self.ingest_api.getSubmissionEnvelope(submissionUrl)
                self.ingest_api.linkEntity(projectIngest, submissionEnvelope, "submissionEnvelopes")
            else:
                projectIngest = {"content" :
                                     {"project_core":
                                     {"project_shortname" : "dummy_project_id"}}}

            projectId = projectIngest["content"]["project_core"]["project_shortname"]

            self.dumpJsonToFile(projectIngest, projectId, "existing_project")

        protocolMap = {}
        for index, protocol in enumerate(protocols):
            if "protocol_id" not in protocol["protocol_core"]:
                raise ValueError('Protocol must have an id attribute')
            self.dumpJsonToFile(protocol, projectId, "protocol_" + str(index))
            protocolMap[protocol["protocol_core"]["protocol_id"]] = protocol
            if not self.dryrun:
                protocolIngest = self.ingest_api.createProtocol(submissionUrl, json.dumps(protocol))
                # self.ingest_api.linkEntity(protocolIngest, projectIngest, "projects")
                protocolMap[protocol["protocol_core"]["protocol_id"]] = protocolIngest
            # else:
            #     linksList.append("protocol_"+protocol["protocol_id"]+"-project_"+projectId)

        biomaterialMap = {}

        # for index, donor in enumerate(donors):
        #     if "sample_id" not in donor:
        #         raise ValueError('Sample of type donor must have an id attribute')
        #     sample_id = donor["sample_id"]
        #
        #     # removing the explicit check for donor absence as this should be picked up by the validator
        #     # if "donor" not in donor:
        #         # Returns ValueError if there are no other donor.fields and donor.is_living is missing
        #         # raise ValueError('Field is_living for sample ' + sample_id + ' is a required field and must either contain one of yes, true, no, or false')
        #     # else:
        #     if "donor" in donor:
        #         if "is_living" in donor["donor"]:
        #             if donor["donor"]["is_living"].lower()in ["true", "yes"]:
        #                 donor["donor"]["is_living"] = True
        #             elif donor["donor"]["is_living"].lower() in ["false", "no"]:
        #                 donor["donor"]["is_living"] = False
        #             '''
        #             # Commented out because we shouldn't be doing content validation in the converter
        #             else:
        #                 # Returns ValueError if donor.is_living isn't true,yes,false,no
        #                 raise ValueError('Field is_living for sample ' + sample_id + ' is a required field and must either contain one of yes, true, no, or false')
        #             '''
        #         # Removing the ValueError if is_living is absent or in the wrong format as this should be flagged by the validator
        #         # else:
        #             # Returns ValueError if there are other donor.fields but donor.is_living is empty
        #             # raise ValueError('Field is_living for sample ' + sample_id + ' is a required field and must either contain one of yes, true, no, or false')
        #
        #     # Removing ValueError for absence of ncbi_taxon_id as this should be caught by the validator
        #     # if "ncbi_taxon_id" not in donor:
        #         # Returns ValueError if donor.ncbi_taxon_id is empty
        #         # raise ValueError('Field ncbi_taxon_id for sample ' + sample_id + ' is a required field and must contain a valid NCBI Taxon ID')
        #
        #     if "ncbi_taxon_id" in donor and  "genus_species" in donor:
        #         donor["genus_species"]["ontology"] = "NCBITaxon:" + str(donor["ncbi_taxon_id"])
        #
        #
        #     sampleMap[sample_id] = donor
        #     donorIds.append(sample_id)
        #
        #     sampleProtocols = []
        #     if "protocol_ids" in donor:
        #         for sampleProtocolId in donor["protocol_ids"]:
        #             if sampleProtocolId not in protocolMap:
        #                 raise ValueError('Sample ' + sample_id
        #                                  + ' references a protocol ' + sampleProtocolId + ' that isn\'t in the protocol worksheet')
        #         sampleProtocols = donor["protocol_ids"]
        #         del donor["protocol_ids"]
        #
        #     donor["core"] = {"type" : "sample",
        #                      "describedBy": self.schema_url + "biomaterial.json",
        #                     "schema_version": self.schema_version}
        #
        #     self.dumpJsonToFile(donor, projectId, "donor_" + str(index))
        #     if not self.dryrun:
        #         biomaterialIngest = self.ingest_api.createBiomaterial(submissionUrl, json.dumps(donor))
        #         self.ingest_api.linkEntity(biomaterialIngest, projectIngest, "projects")
        #         biomaterialMap[biomaterial_id] = biomaterialIngest
        #
        #         if biomaterialProtocols:
        #             for biomaterialProtocolId in biomaterialProtocols:
        #                 self.ingest_api.linkEntity(biomaterialIngest, protocolMap[biomaterialProtocolId], "protocols")
        #     else:
        #         linksList.append("biomaterial_" + str(biomaterial_id) + "-project_" + str(projectId))
        #         if biomaterialProtocols:
        #             for biomaterialProtocolId in biomaterialProtocols:
        #                 linksList.append("biomaterial_" + str(biomaterial_id) + "-protocol_" + str(biomaterialProtocolId))


        for index, biomaterial in enumerate(biomaterials):
            if "biomaterial_id" not in biomaterial["biomaterial_core"]:
                raise ValueError('Biomaterial must have an id attribute')
            biomaterialMap[biomaterial["biomaterial_core"]["biomaterial_id"]] = biomaterial
            biomaterial_id = biomaterial["biomaterial_core"]["biomaterial_id"]

            # if "ncbi_taxon_id" not in biomaterial:
                # Returns ValueError if donor.ncbi_taxon_id is empty
                # raise ValueError(
                #     'Field ncbi_taxon_id for biomaterial ' + biomaterial_id + ' is a required field and must contain a valid NCBI Taxon ID')

            if "ncbi_taxon_id" in biomaterial and "genus_species" in biomaterial:
                biomaterial["genus_species"]["ontology"] = "NCBITaxon:" + str(biomaterial["ncbi_taxon_id"])

        # add dependent information to various biomaterial types

        for publication in cell_line_publications:
            if "biomaterial_id" in publication["biomaterial_core"]:
                bio_id = publication["biomaterial_core"]["biomaterial_id"]
                del publication["biomaterial_core"]

                if "publications" not in biomaterialMap[bio_id]:
                    biomaterialMap[bio_id]["publications"] = []
                biomaterialMap[bio_id]["publications"].append(publication)

        for familialRel in familialRelationships:
            if "biomaterial_id" in familialRel["biomaterial_core"]:
                bio_id = familialRel["biomaterial_core"]["biomaterial_id"]
                del familialRel["biomaterial_core"]

                if "familial_relationship" not in biomaterialMap[bio_id]:
                    biomaterialMap[bio_id]["familial_relationship"] = []
                biomaterialMap[bio_id]["familial_relationship"].append(familialRel)

        # build the process map from the different types of process infromation
        processMap = {}

        for index, process in enumerate(processes):
            if "process_id" not in process["process_core"]:
                raise ValueError('Process must have an id attribute')
            processMap[process["process_core"]["process_id"]] = process

        for reagent in reagents:
            if "process_id" in reagent["process_core"]:
                process_id = reagent["process_core"]["process_id"]
                del reagent["process_core"]

                if "process_reagents" not in processMap[process_id]:
                    processMap[process_id]["process_reagents"] = []
                processMap[process_id]["process_reagents"].append(reagents)

        # submit biomaterials to ingest and link to project and protocols
        for index, biomaterial_id in enumerate(biomaterialMap.keys()):
            biomaterial = biomaterialMap[biomaterial_id]
            if "has_input_biomaterial" in biomaterial["biomaterial_core"]:
                if biomaterial["biomaterial_core"]["has_input_biomaterial"] not in biomaterialMap.keys():
                    raise ValueError('Biomaterial '+ str(biomaterial_id) +' references another biomaterial '+ str(biomaterial["biomaterial_core"]["has_input_biomaterial"]) +' that isn\'t in the spraedsheet')


            if "process_ids" in biomaterial:
                for process_id in biomaterial["process_ids"]:
                    if process_id not in processMap.keys():
                        raise ValueError(
                         'A biomaterial references a process ' + process_id + ' that isn\'t in the biomaterials worksheet')
                    if "biomaterial_ids" not in processMap[process_id]:
                        processMap[process_id]["biomaterial_ids"] = []
                    processMap[process_id]["biomaterial_ids"].append(biomaterial["biomaterial_core"]["biomaterial_id"])
                del biomaterial["process_ids"]


            self.dumpJsonToFile(biomaterial, projectId, "biomaterial_" + str(index))
            if not self.dryrun:
                biomaterialIngest = self.ingest_api.createSample(submissionUrl, json.dumps(biomaterial))
                self.ingest_api.linkEntity(biomaterialIngest, projectIngest, "projects")
                biomaterialMap[biomaterial["biomaterial_core"]["biomaterial_id"]] = biomaterialIngest
                # if biomaterialProtocols:
                #     for biomaterialProtocolId in biomaterialProtocols:
                #         self.ingest_api.linkEntity(biomaterialIngest, protocolMap[biomaterialProtocolId], "protocols")
            else:
                linksList.append("biomaterial_" + str(biomaterial_id) + "-project_" + str(projectId))
                # if biomaterialProtocols:
                #     for biomaterialProtocolId in biomaterialProtocols:
                #         linksList.append("biomaterial_" + str(biomaterial_id) + "-protocol_" + str(biomaterialProtocolId))

        # create has_input_biomaterial links between biomaterials separately to make sure all biomaterials are submitted
        for index, biomaterial_id in enumerate(biomaterialMap.keys()):
            if not self.dryrun:
                if "has_input_biomaterial" in biomaterialMap[biomaterial_id]['content']["biomaterial_core"]:
                    self.ingest_api.linkEntity(biomaterialMap[biomaterial_id],
                                               biomaterialMap[biomaterialMap[biomaterial_id]['content']["biomaterial_core"]["has_input_biomaterial"]],
                                               "hasInputBiomaterial")

            else:
                if "has_input_biomaterial" in biomaterialMap[biomaterial_id]:
                    linksList.append(
                        "biomaterial_" + str(biomaterial_id) + "-hasInputBiomaterial_" + str(biomaterialMap[biomaterial_id]["biomaterial_core"]["has_input_biomaterial"]))



        filesMap={}
        for index, file in enumerate(files):
            if "file_name" not in file["file_core"]:
                raise ValueError('Files must have a name')
            if "process_id" not in file:
                raise ValueError('Files must be linked to a process')
            process = file["process_id"]
            biomaterial = file["biomaterial_id"]
            del file["process_id"]
            del file["biomaterial_id"]
            filesMap[file["file_core"]["file_name"]] = file

            if "files" not in processMap[process]:
                processMap[process]["files"] = []

            # ????do we need this????
            processMap[process]["files"].append(file["file_core"]["file_name"])

            self.dumpJsonToFile(file, projectId, "files_" + str(index))
            if not self.dryrun:
                fileIngest = self.ingest_api.createFile(submissionUrl, file["file_core"]["file_name"], json.dumps(file))
                filesMap[file["file_core"]["file_name"]] = fileIngest

            #     if biomaterial in biomaterialMap:
            #         self.ingest_api.linkEntity(fileIngest, biomaterialMap[biomaterial], "biomaterials")
            # else:
            #     if biomaterial in biomaterialMap:
            #         linksList.append("file_" + file["file_name"] + "-biomaterial_" + biomaterial)

        for index, process in enumerate(processMap.values()):
            if "process_id" not in process["process_core"]:
                raise ValueError('Each process must have an id attribute' + str(process))
            # if "files" not in process:
                # raise ValueError('Each process must list associated files using the files attribute')
            # else:
            if "files" in process:
                for file in process["files"]:
                    if file not in filesMap:
                        raise ValueError('Process references file '+file+' that isn\'t defined in the files sheet')
                files = process["files"]
                del process["files"]

            if "biomaterial_ids" not in process:
                raise ValueError("Every process must reference a biomaterial using the biomaterial_id attribute")
            else:
                for biomaterial_id in process["biomaterial_ids"]:
                    if biomaterial_id not in biomaterialMap:
                        raise ValueError('An process references a biomaterial '+biomaterial_id+' that isn\'t in one of the biomaterials worksheets')
            biomaterials = process["biomaterial_ids"]
            del process["biomaterial_ids"]

            self.dumpJsonToFile(process, projectId, "process_" + str(index))
            # ???this is the bit we still need to figure out????
            if not self.dryrun:
                processIngest = self.ingest_api.createAssay(submissionUrl, json.dumps(process))
                self.ingest_api.linkEntity(processIngest, projectIngest, "projects")

                if biomaterials in biomaterialMap:
                    self.ingest_api.linkEntity(processIngest, biomaterialMap[biomaterials], "biomaterials")

                for file in files:
                    self.ingest_api.linkEntity(processIngest, filesMap[file], "files")
            else:
                linksList.append("process_" + str(process["process_core"]["process_id"]) + "-project_" + str(projectId))

                for biomaterial in biomaterials:
                    linksList.append("process_" + str(process["process_core"]["process_id"]) + "-biomaterial_" + str(biomaterial))

                for file in files:
                    linksList.append("process_" + str(process["process_core"]["process_id"]) + "-file_" + str(file))

        self.dumpJsonToFile(linksList, projectId, "dry_run_links")
        self.logger.info("All done!")
        wb.close()
        return submissionUrl

if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option("-p", "--path", dest="path",
                      help="path to HCA example data bundles", metavar="FILE")
    parser.add_option("-d", "--dry", help="doa dry run without submitting to ingest", action="store_true", default=False)
    parser.add_option("-o", "--output", dest="output",
                      help="output directory where to dump json files submitted to ingest", metavar="FILE", default=None)
    parser.add_option("-i", "--id", dest="project_id",
                      help="The project_id for an existing submission", default=None)
    parser.add_option("-v", "--version", dest="schema_version", help="Metadata schema version", default=None)


    (options, args) = parser.parse_args()
    if not options.path:
        print ("You must supply path to the HCA bundles directory")
        exit(2)
    submission = SpreadsheetSubmission(options.dry, options.output, options.schema_version)
    submission.submit(options.path, None, None, options.project_id)
