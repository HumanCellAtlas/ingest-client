import logging

import openpyxl
from openpyxl import load_workbook

from ingest.importer.conversion import template_manager
from ingest.importer.conversion.metadata_entity import MetadataEntity
from ingest.importer.conversion.template_manager import TemplateManager
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook
from ingest.importer.spreadsheet.ingest_worksheet import IngestWorksheet
from ingest.importer.submission import IngestSubmitter, EntityMap, EntityLinker
from ingest.utils.IngestError import ImporterError, ParserError

format = '[%(filename)s:%(lineno)s - %(funcName)20s() ] %(asctime)s - %(name)s - %(levelname)s - %(message)s'
logging.basicConfig(format=format)


class XlsImporter:

    # TODO why does the importer need to refer to an IngestApi instance?
    # Seems like it should be the IngestSubmitter that takes care of this detail
    # alegria: Submitter is part of the Ingest Importer
    def __init__(self, ingest_api):
        self.ingest_api = ingest_api
        self.logger = logging.getLogger(__name__)

    def dry_run_import_file(self, file_path, project_uuid=None):
        spreadsheet_json, template_mgr, errors = self._generate_spreadsheet_json(file_path, project_uuid)
        entity_map = self._process_links_from_spreadsheet(template_mgr, spreadsheet_json)

        return entity_map

    def _generate_spreadsheet_json(self, file_path, project_uuid=None):
        ingest_workbook = self._create_ingest_workbook(file_path)

        try:
            template_mgr = template_manager.build(ingest_workbook.get_schemas(), self.ingest_api)
        except Exception as e:
            raise SchemaRetrievalError(
                f'There was an error retrieving the schema information to process the spreadsheet. {str(e)}')

        workbook_importer = WorkbookImporter(template_mgr)
        spreadsheet_json, errors = workbook_importer.do_import(ingest_workbook, project_uuid)

        return spreadsheet_json, template_mgr, errors

    def import_file(self, file_path, submission_url, project_uuid=None):
        submission = None
        try:
            spreadsheet_json, template_mgr, errors = self._generate_spreadsheet_json(file_path, project_uuid)
            if not errors:
                entity_map = self._process_links_from_spreadsheet(template_mgr, spreadsheet_json)

                # TODO the submission_url should be passed to the IngestSubmitter instead
                submitter = IngestSubmitter(self.ingest_api)
                submission = submitter.submit(entity_map, submission_url)
                self.logger.info(f'Submission in {submission_url} is done!')
            else:
                self.report_errors(submission_url, errors)
        except Exception as e:
            self.ingest_api.create_submission_error(submission_url, ImporterError(str(e)).getJSON())
            self.logger.error(str(e), exc_info=True)
        return submission

    def report_errors(self, submission_url, errors):
        for error in errors:
            self.ingest_api.create_submission_error(
                submission_url,
                ParserError(error["location"], error["type"], error["detail"]).getJSON()
            )
            self.logger.error(error["detail"], exc_info=True)

    def insert_uuids(self, submission, file_path):
        if not submission:
            return

        wb = load_workbook(filename=file_path)

        worksheets = {}
        col_idx = 1
        for entity in submission.get_entities():
            if entity.spreadsheet_location:
                worksheet_title = entity.spreadsheet_location.get('worksheet_title')
                row_index = entity.spreadsheet_location.get('row_index')

                if not worksheets.get(worksheet_title):
                    worksheet = wb.get_sheet_by_name(worksheet_title)
                    ingest_worksheet = IngestWorksheet(worksheet=worksheet)
                    column_header = f'{entity.concrete_type}.uuid'
                    if column_header in ingest_worksheet.get_column_headers():
                        continue
                    ingest_worksheet.insert_column_with_header(column_header, col_idx)
                    worksheets[worksheet_title] = ingest_worksheet

                ingest_worksheet = worksheets.get(worksheet_title)
                ingest_worksheet.cell(row=row_index, column=col_idx).value = entity.uuid

        return wb.save(file_path)

    @staticmethod
    def _create_ingest_workbook(file_path):
        workbook = openpyxl.load_workbook(filename=file_path, read_only=True)
        return IngestWorkbook(workbook)

    @staticmethod
    def _process_links_from_spreadsheet(template_mgr, spreadsheet_json):
        entity_map = EntityMap.load(spreadsheet_json)
        entity_linker = EntityLinker(template_mgr)
        entity_map = entity_linker.process_links_from_spreadsheet(entity_map)
        return entity_map


_PROJECT_ID = 'project_0'
_PROJECT_TYPE = 'project'


class _ImportRegistry:
    """
    This is a helper class for managing metadata entities during Workbook import.
    """

    def __init__(self):
        self._submittable_registry = {}
        self._module_registry = {}
        self._module_list = []
        self.project_id = _PROJECT_ID

    def add_submittable(self, metadata: MetadataEntity):
        # TODO no test to check case sensitivity
        domain_type = metadata.domain_type.lower()
        type_map = self._submittable_registry.get(domain_type)
        if not type_map:
            type_map = {}
            self._submittable_registry[domain_type] = type_map
        if domain_type.lower() == _PROJECT_TYPE:
            if not type_map.get(self.project_id):
                metadata.object_id = metadata.object_id or self.project_id
                self.project_id = metadata.object_id
            else:
                raise MultipleProjectsFound()
        type_map[metadata.object_id] = metadata

    def add_module(self, metadata: MetadataEntity):
        if metadata.domain_type.lower() == 'project':
            metadata.object_id = self.project_id
        self._module_list.append(metadata)

    def import_modules(self):
        for module_entity in self._module_list:
            type_map = self._submittable_registry.get(module_entity.domain_type)
            submittable_entity = type_map.get(module_entity.object_id)
            submittable_entity.add_module_entity(module_entity)

    def flatten(self):
        flat_map = {}
        for domain_type, type_map in self._submittable_registry.items():
            flat_type_map = {object_id: metadata.map_for_submission()
                             for object_id, metadata in type_map.items()}
            flat_map[domain_type] = flat_type_map
        return flat_map

    def has_project(self):
        project_registry = self._submittable_registry.get(_PROJECT_TYPE)
        return project_registry and project_registry.get(self.project_id)


class WorkbookImporter:

    def __init__(self, template_mgr):
        self.worksheet_importer = WorksheetImporter(template_mgr)
        self.template_mgr = template_mgr
        self.logger = logging.getLogger(__name__)

    def do_import(self, workbook: IngestWorkbook, project_uuid=None):
        registry = _ImportRegistry()
        importable_worksheets = workbook.importable_worksheets()

        if project_uuid:
            project_metadata = MetadataEntity(domain_type=_PROJECT_TYPE,
                                              concrete_type=_PROJECT_TYPE,
                                              object_id=project_uuid,
                                              is_reference=True,
                                              content={})
            registry.add_submittable(project_metadata)

            importable_worksheets = [ws for ws in importable_worksheets
                                     if _PROJECT_TYPE not in ws.title.lower()]
        workbook_errors = []
        for worksheet in importable_worksheets:
            try:
                metadata_entities, worksheet_errors = self.worksheet_importer.do_import(worksheet)
                module_field_name = worksheet.get_module_field_name()
                for error in worksheet_errors:
                    if error["location"]:
                        error["location"] = f'sheet={worksheet.title}, {error["location"]}'
                    else:
                        error["location"] = f'sheet={worksheet.title}'
                    workbook_errors.append(error)

                for entity in metadata_entities:
                    if worksheet.is_module_tab():
                        entity.retain_content_fields(module_field_name)
                        registry.add_module(entity)
                    else:
                        registry.add_submittable(entity)
            except Exception as e:
                workbook_errors.append({"location": "File", "type": e.__class__.__name__, "detail": str(e)})

        if registry.has_project():
            registry.import_modules()
        else:
            raise NoProjectFound()
        return registry.flatten(), workbook_errors


class WorksheetImporter:
    KEY_HEADER_ROW_IDX = 4
    USER_FRIENDLY_HEADER_ROW_IDX = 2
    START_ROW_OFFSET = 5

    UNKNOWN_ID_PREFIX = '_unknown_'

    def __init__(self, template: TemplateManager):
        self.template = template
        self.unknown_id_ctr = 0
        self.logger = logging.getLogger(__name__)
        self.concrete_entity = None

    def do_import(self, ingest_worksheet: IngestWorksheet):
        records = []
        worksheet_errors = []
        try:
            row_template = self.template.create_row_template(ingest_worksheet)
            rows = ingest_worksheet.get_data_rows()
            for index, row in enumerate(rows):
                metadata, row_errors = row_template.do_import(row)
                for error in row_errors:
                    if error["location"]:
                        error["location"] = f'row={index}, {error["location"]}'
                    else:
                        error["location"] = f'row={index}'
                    worksheet_errors.append(error)
                if not metadata.object_id:
                    metadata.object_id = self._generate_id()
                records.append(metadata)
        except Exception as e:
            worksheet_errors.append({"location": "", "type": e.__class__.__name__, "detail": str(e)})
        return records, worksheet_errors

    def _generate_id(self):
        self.unknown_id_ctr = self.unknown_id_ctr + 1
        return f'{self.UNKNOWN_ID_PREFIX}{self.unknown_id_ctr}'


class MultipleProjectsFound(Exception):
    def __init__(self):
        message = f'The spreadsheet should only be associated to a single project.'
        super(MultipleProjectsFound, self).__init__(message)


class NoProjectFound(Exception):
    def __init__(self):
        message = f'The spreadsheet should be associated to a project.'
        super(NoProjectFound, self).__init__(message)


class SchemaRetrievalError(Exception):
    pass
