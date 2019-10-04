from openpyxl import Workbook, load_workbook

from ingest.importer.spreadsheet.ingest_worksheet import IngestWorksheet
from ingest.importer.submission import Submission

SCHEMAS_WORKSHEET = 'Schemas'
SPECIAL_TABS = [SCHEMAS_WORKSHEET]


class IngestWorkbook:

    def __init__(self, workbook: Workbook):
        self.workbook = workbook

    @classmethod
    def from_file(cls, file_path, read_only=True) -> 'IngestWorkbook':
        workbook = load_workbook(filename=file_path, read_only=read_only)
        return cls(workbook)

    def get_worksheet(self, worksheet_title):
        if worksheet_title in self.workbook.sheetnames:
            return self.workbook[worksheet_title]

        if worksheet_title.lower() in self.workbook.sheetnames:
            return self.workbook[worksheet_title.lower()]

        return None

    def get_schemas(self):
        schemas = []

        worksheet = None

        if SCHEMAS_WORKSHEET in self.workbook.sheetnames:
            worksheet = self.workbook[SCHEMAS_WORKSHEET]

        if SCHEMAS_WORKSHEET.lower() in self.workbook.sheetnames:
            worksheet = self.workbook[SCHEMAS_WORKSHEET.lower()]

        if not worksheet:
            return schemas

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            schema_cell = row[0]
            schemas.append(schema_cell.value)
        return schemas

    def importable_worksheets(self):
        return [IngestWorksheet(worksheet) for worksheet in self.workbook.worksheets
                if worksheet.title not in SPECIAL_TABS]

    def add_entity_uuids(self, submission: Submission):
        worksheets = {}
        col_idx = 1
        entities = [entity for entity in submission.get_entities() if entity.spreadsheet_location]
        for entity in entities:
            worksheet_title = entity.spreadsheet_location.get('worksheet_title')
            row_index = entity.spreadsheet_location.get('row_index')

            if not worksheets.get(worksheet_title):
                worksheet = self.workbook[worksheet_title]
                ingest_worksheet = IngestWorksheet(worksheet=worksheet)
                worksheets[worksheet_title] = ingest_worksheet

            column_header = f'{entity.concrete_type}.uuid'

            if column_header not in ingest_worksheet.get_column_headers():
                ingest_worksheet.insert_column_with_header(column_header, col_idx)

            ingest_worksheet = worksheets.get(worksheet_title)
            ingest_worksheet.cell(row=row_index, column=col_idx).value = entity.uuid

    def add_schemas_worksheet(self, schemas):
        if SCHEMAS_WORKSHEET not in self.workbook.sheetnames:
            schema_sheet = self.workbook.create_sheet(SCHEMAS_WORKSHEET)
            schema_sheet.insert_cols(1)
            row_idx = 1
            col_idx = 1
            schema_sheet.cell(row=row_idx, column=col_idx).value = SCHEMAS_WORKSHEET
            for schema in schemas:
                row_idx += 1
                schema_sheet.cell(row=row_idx, column=col_idx).value = schema

    def save(self, file_path):
        self.workbook.save(file_path)
