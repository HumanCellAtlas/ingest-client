#!/usr/bin/env python
"""
This script will read a spreadsheet, generate a manifest, submit all items to the ingest API, 
assign uuid and generate a directory of bundles for the submitted data
"""
from openpyxl.utils.exceptions import InvalidFileException

__author__ = "jupp"
__license__ = "Apache 2.0"

import json, os
import logging
import ingest.utils.token_util as token_util


from openpyxl import load_workbook
from ingest.api.ingestapi import IngestApi

from optparse import OptionParser

from itertools import chain
from collections import defaultdict


# these are spreadsheet fields that can be a list
# todo - these should be read out of the json schema at the start

schema_arrayFields = {
    "cell_line": ["genus_species", "publications", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "cell_suspension": ["genus_species", "target_cell_type", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "donor_organism": ["genus_species", "disease", "familial_relationship", "ethnicity", "strain", "ncbi_taxon_id",
                       "supplementary_files", "process_ids"],
    "organoid": ["genus_species", "ncbi_taxon_id", "supplementary_files", "process_ids"],
    "specimen_from_organism": ["genus_species", "disease", "ncbi_taxon_id", "supplementary_files", "process_ids",
                               "gross_image", "microscopic_image"],
    "sequence_file": ["insdc_run"],
    "analysis_process": ["inputs", "tasks", "input_bundles", "outputs", "operator_identity", "protocol_ids"],
    "collection_process": ["process_reagents", "operator_identity", "protocol_ids"],
    "dissociation_process": ["process_reagents", "operator_identity", "protocol_ids"],
    "imaging_process": ["field_counts", "field_microns", "field_resolution", "operator_identity", "protocol_ids"],
    "enrichment_process": ["process_reagents", "operator_identity", "protocol_ids"],
    "library_preparation_process": ["process_reagents", "operator_identity", "protocol_ids"],
    "sequencing_process": ["process_reagents", "operator_identity", "protocol_ids"],
    "project": ["contributors", "supplementary_files", "publications", "authors"],
    "publication": ["authors"]
}

schema_timeFields = {
    "cell_line": ["date_established"],
    "donor_organism": ["time_of_death"],
    "specimen_from_organism": ["collection_time"],
    "analysis_process": ["start_time", "stop_time", "timestamp_start_utc", "timestamp_stop_utc"],
    "collection_process": ["start_time"],
    "dissociation_process": ["start_time"],
    "enrichment_process": ["start_time"],
    "imaging_process": ["start_time"],
    "library_preparation_process": ["start_time"],
    "sequencing_process": ["start_time"]
}

schema_booleanFields = {
    "donor_organism": ["is_living", "cold_perfused"],
    "sequencing_process": ["paired_ends"],
    "death": ["cold_perfused"]
}

schema_integerFields = {
    "cell_line": ["ncbi_taxon_id", "passage_number"],
    "cell_suspension": ["total_estimated_cells", "ncbi_taxon_id", "passage_number"],
    "donor_organism": ["ncbi_taxon_id", "hardy_scale"],
    "organoid": ["ncbi_taxon_id", "passage_number"],
    "specimen_from_organism": ["ncbi_taxon_id", "ischemic_time", "postmortem_interval"],
    "sequence_file": ["lane_index", "read_length"],
    "analysis_process": ["cpus"],
    "imaging_process": ["field_counts", "field_microns", "field_resolution"],
    "library_preparation_process": ["spike_in_dilution", "barcode_offset", "barcode_length"],
    "project": ["pmid"],
    "project.publications": ["pmid"]
}

schema_numberFields = {
    "cell_morphology": ["cell_size", "cell_viability", "percent_necrosis"],
    "imaging_process": ["exposure_time"],
    "enrichment_process": ["min_size_selected", "max_size_selected"],
    "organoid": ["organoid_age", "cell_size", "cell_viability", "percent_necrosis"],
    "cell_line": ["cell_size", "cell_viability", "percent_necrosis"],
    "cell_suspension": ["cell_size", "cell_viability", "percent_necrosis"],
    "donor_organism": ["days_on_ventilator", "body_mass_index"],
    "specimen_from_organism": ["storage_time"]
}

# maps sheets to the latest version of each schema
# todo - this should be replaced by dynamic lookups against ingest-core /schemas endpoint
# (spec for lookup: query by 'title' (sheet name in the spreadsheet, field in the schema)
schema_sheetname_mappings = {
    # biomaterials
    "cell_line": "https://schema.humancellatlas.org/type/biomaterial/5.1.0/cell_line",
    "cell_suspension": "https://schema.humancellatlas.org/type/biomaterial/5.1.0/cell_suspension",
    "donor_organism": "https://schema.humancellatlas.org/type/biomaterial/5.1.0/donor_organism",
    "organoid": "https://schema.humancellatlas.org/type/biomaterial/5.1.0/organoid",
    "specimen_from_organism": "https://schema.humancellatlas.org/type/biomaterial/5.1.0/specimen_from_organism",
    # files
    "analysis_file": "https://schema.humancellatlas.org/type/file/5.1.0/analysis_file",
    "sequence_file": "https://schema.humancellatlas.org/type/file/5.1.0/sequence_file",
    # analysis processes
    "analysis_process": "https://schema.humancellatlas.org/type/process/analysis/5.1.0/analysis_process",
    # biomaterial processes
    "process": "https://schema.humancellatlas.org/type/process/1.0.0/process",
    "collection_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.1.0/collection_process",
    "dissociation_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.1.0/dissociation_process",
    "enrichment_process": "https://schema.humancellatlas.org/type/process/biomaterial_collection/5.1.0/enrichment_process",
    # imaging processes
    "imaging_process": "https://schema.humancellatlas.org/type/process/imaging/5.1.0/imaging_process",
    # sequencing processes
    "library_preparation_process": "https://schema.humancellatlas.org/type/process/sequencing/5.1.0/library_preparation_process",
    "sequencing_process": "https://schema.humancellatlas.org/type/process/sequencing/5.1.0/sequencing_process",
    # projects
    "project": "https://schema.humancellatlas.org/type/project/5.1.0/project",
    # protocols
    "protocol": "https://schema.humancellatlas.org/type/protocol/5.1.0/protocol",
    "analysis_protocol": "https://schema.humancellatlas.org/type/protocol/analysis/5.1.0/analysis_protocol",
    "biomaterial_collection_protocol": "https://schema.humancellatlas.org/type/protocol/biomaterial/5.1.0/biomaterial_collection_protocol",
    "imaging_protocol": "https://schema.humancellatlas.org/type/protocol/imaging/5.1.0/imaging_protocol",
    "sequencing_protocol": "https://schema.humancellatlas.org/type/protocol/sequencing/5.1.0/sequencing_protocol"
}


class SpreadsheetSubmission:

    def __init__(self, dry=False, output=None, schema_version=None):
        # todo - the logging code below doesn't work in python 3 - we should upgrade this
        formatter = logging.Formatter(
            '[%(filename)s:%(lineno)s - %(funcName)20s() ] %(asctime)s - %(name)s - %(levelname)s - %(message)s')
        logging.basicConfig(formatter=formatter, level=logging.INFO)
        self.logger = logging.getLogger(__name__)

        self.dryrun = dry
        self.outputDir = output
        self.ingest_api = None
        if not self.dryrun:
            self.ingest_api = IngestApi()

    def createSubmission(self, token):
        self.logger.info("creating submission...")
        if not self.ingest_api:
            self.ingest_api = IngestApi()

        submissionUrl = self.ingest_api.createSubmission(token)
        self.logger.info("new submission " + submissionUrl)
        return submissionUrl

    # this only works for keys nested to two levels, need something smarter to handle arbitrary
    # depth keys e.g. we support <level1>.<level2> = <value>, where value is either a single value or list
    # this doesn't support <level1>.<level2>.<level3> except if <level3> is "ontology"
    #This function takes a dictionary object, a key/value pair from the spreadsheet input and an object type, and adds
    #the key/value pair to the dictionary. The type is used to assess whether a field is an ontology or an array based on the
    # two lookup dictionaries declared at the top.
    # For || separated strings, the value field is split. For json arrays, the value is put into an array. Json objects of type
    # ontology require some extra formatting, which is also done here.

    def _keyValueToNestedObject(self, obj, key, value, type):
        if "*" in key:
            key = key.replace("*", "")
        d = str(value)
        # If the value contains a double pipe (||) or the key is for a field that can be a list (with or without also being
        # an ontology field), put value into an array (splitting if necessary)
        if "\"" in str(value) or "||" in str(value) \
                or (type in schema_arrayFields.keys() and key.split('.')[-1] in schema_arrayFields[type]) \
                or (type in schema_arrayFields.keys() and (key.split('.')[-1] == "ontology" or key.split('.')[-1] == "text")
                    and key.split('.')[-2] in schema_arrayFields[type]):
            if "||" in str(value):
                d = str(value).split("||")
            else:
                d = [value]

            if key.split('.')[-1] in ["text", "ontology"]:

                if isinstance(d, list):
                    t = []
                    for index, v in enumerate(d):
                        t.append({key.split('.')[-1] : d[index]})
                    d = t
                else:
                    d = {key.split('.')[-1] : d}
                key = ".".join(key.split('.')[:-1])

        # Raise an error if the key is too nested
        if len(key.split('.')) > 3:
            raise ValueError('We don\'t support keys nested greater than 3 levels, found:'+key)

        # If the key is in the date_time field list, convert the date time into a string of format YYYY-MM-DDThh:mm:ssZ
        # so it validates. Same principle for integers and boolean types
        if type in schema_timeFields.keys():
            try:
                if key.split('.')[-1] in schema_timeFields[type]:
                    if isinstance(d, list):
                        for i, v in enumerate(d):
                            date_string = v.strftime("%Y-%m-%dT%H:%M:%SZ")
                            d[i] = date_string
                    else:
                        d = d.strftime("%Y-%m-%dT%H:%M:%SZ")
            except:
                self.logger.warn('Failed to convert field {0} in sheet {1} (value {2}) to date_time value'.format(key, type, d))
                d = str(d)
        if type in schema_integerFields.keys():
            try:
                if key.split('.')[-1] in schema_integerFields[type]:
                    if isinstance(d, list):
                        for i, v in enumerate(d):
                            d[i] = int(v)
                    else:
                        d = int(d)
            except:
                self.logger.warn('Failed to convert field {0} in sheet {1} (value {2}) to integer value'.format(key, type, d))
                d = str(d)
        if type in schema_numberFields.keys():
            try:
                if key.split('.')[-1] in schema_numberFields[type]:
                    if isinstance(d, list):
                        for i, v in enumerate(d):
                            d[i] = float(v)
                    else:
                        d = float(d)
            except:
                self.logger.warn(
                    'Failed to convert field {0} in sheet {1} (value {2}) to float value'.format(key, type, d))
                d = str(d)
        if type in schema_booleanFields.keys():
            try:
                if key.split('.')[-1] in schema_booleanFields[type]:
                    if isinstance(d, list):
                        for i, v in enumerate(d):
                            if v.lower() in ["true", "yes"]:
                                d[i] = True
                            elif v.lower() in ["false", "no"]:
                                d[i] = False
                    else:
                        if d.lower() in ["true", "yes"]:
                            d = True
                        elif d.lower() in ["false", "no"]:
                            d = False
            except:
                self.logger.warn('Failed to convert field {0} in sheet {1} (value {2}) to boolean value'.format(key, type, d))
                d = str(d)

        # Build up the key-value dictionary
        for part in reversed(key.split('.')):
            d = {part: d}

        return self._mergeDict(obj, d)

    def _mergeDict(self, dict1, dict2):
        dict3 = defaultdict(list)
        for k, v in chain(dict1.items(), dict2.items()):
            if k in dict3:
                if isinstance(v, dict):
                    dict3[k].update(self._mergeDict(dict3[k], v))
                elif isinstance(v, list) and isinstance(dict3[k], list) and len(v) == len(dict3[k]):
                    for index, e in enumerate(v):
                        dict3[k][index].update(self._mergeDict(dict3[k][index], e))
                else:
                    dict3[k].update(v)
            else:
                dict3[k] = v
        return dict3

    #sheets with one or more data rows and properties in row 1
    def _multiRowToObjectFromSheet(self, type, sheet):
        objs = []
        for row in sheet.iter_rows(row_offset=3, max_row=(sheet.max_row - 3)):
            obj = {}
            hasData = False
            for cell in row:
                # WARNING: remove long from the checks if using python 3!
                if not cell.value and not isinstance(cell.value, (int, float)):
                    continue
                hasData = True
                cellCol = cell.col_idx
                propertyValue = sheet.cell(row=3, column=cellCol).value

                d = self._keyValueToNestedObject(obj, propertyValue, cell.value, type)
                obj.update(d)
            if hasData:
                self.logger.debug(json.dumps(obj))
                objs.append(obj)

        return objs

    # sheets that represent one entity where the properties are in column 0
    def _sheetToObject(self, type, sheet):
        obj = {}
        for row in sheet.iter_rows():
            if len(row) > 1:
                propertyCell = row[0].value
                valueCell = row[1].value
                if valueCell:
                    obj = self._keyValueToNestedObject(obj, propertyCell, valueCell, type)
                    # obj.update(d)
        self.logger.debug(json.dumps(obj))
        return obj

    def _emptyProcessObject(self, type, empty_process_id):
        obj = {}
        process_core = {"process_id": str(type) + "_process_" + str(empty_process_id)}
        schema_type = "process"
        describedBy = schema_sheetname_mappings["process"]
        obj["process_core"] = process_core
        obj["schema_type"] = schema_type
        obj["describedBy"] = describedBy
        return obj

    def completeSubmission(self):
        self.ingest_api.finishSubmission()

    def submit(self, pathToSpreadsheet, submissionUrl, token=None, project_id=None):
        try:
            self._process(pathToSpreadsheet, submissionUrl, token, project_id)
        except Exception as e:
            self.logger.error("Error:"+str(e))
            self.logger.error("Exception occurred processing spreadsheet", exc_info=e)
            raise e

    def dumpJsonToFile(self, object, projectId, name):
        if self.outputDir:
            dir = os.path.abspath(self.outputDir)
            if not os.path.exists(dir):
                os.makedirs(dir)
            tmpFile = open(dir + "/" + str(projectId) + "_" + str(name) + ".json", "w")
            tmpFile.write(json.dumps(object, indent=4))
            tmpFile.close()

    def _process(self, pathToSpreadsheet, submissionUrl, token, existing_project_id):

        # parse the spreadsheet
        try:
            wb = load_workbook(filename=pathToSpreadsheet)
        except InvalidFileException:
            raise SpreadsheetUploadError(400, "The uploaded file is not a valid XLSX spreadsheet", "")
        # This code section deals with cases where the project has already been submitted
        # ASSUMPTION: now additional project information (publications, contributors etc) is added via
        # the spreadsheet if the project already exists

        projectSheet = wb.create_sheet()
        projectPubsSheet = wb.create_sheet()
        contributorSheet = wb.create_sheet()
        empty_wrapper_id = 1
        empty_sampling_id = 1

        if existing_project_id is None:
            projectSheet = wb.get_sheet_by_name("project")

            if "project.publications" in wb.sheetnames:
                projectPubsSheet = wb.get_sheet_by_name("project.publications")
            if "contact" in wb.sheetnames:
                contributorSheet = wb.get_sheet_by_name("contact")

        specimenSheet = wb.create_sheet()
        donorSheet = wb.create_sheet()
        cellSuspensionSheet = wb.create_sheet()
        familialRelationshipSheet = wb.create_sheet()

        if "specimen_from_organism" in wb.sheetnames:
            specimenSheet = wb.get_sheet_by_name("specimen_from_organism")
        if "donor_organism" in wb.sheetnames:
            donorSheet = wb.get_sheet_by_name("donor_organism")
        if "cell_suspension" in wb.sheetnames:
            cellSuspensionSheet = wb.get_sheet_by_name("cell_suspension")
        if "familial_relationship" in wb.sheetnames:
            familialRelationshipSheet = wb.get_sheet_by_name("familial_relationship")

        organoidSheet = wb.create_sheet()
        if "organoid" in wb.sheetnames:
            organoidSheet = wb.get_sheet_by_name("organoid")

        clSheet = wb.create_sheet()
        if "cell_line" in wb.sheetnames:
            clSheet = wb.get_sheet_by_name("cell_line")

        clPublicationSheet = wb.create_sheet()
        if "cell_line.publications" in wb.sheetnames:
            clPublicationSheet = wb.get_sheet_by_name("cell_line.publications")


        protocolSheet = wb.create_sheet()
        collectionSheet = wb.create_sheet()
        dissociationSheet = wb.create_sheet()
        enrichmentSheet = wb.create_sheet()
        libraryPrepSheet = wb.create_sheet()
        sequencingSheet = wb.create_sheet()
        reagentsSheet = wb.create_sheet()
        filesSheet = wb.create_sheet()

        if "protocol" in wb.sheetnames:
            protocolSheet = wb.get_sheet_by_name("protocol")
        if "enrichment_process" in wb.sheetnames:
            enrichmentSheet = wb.get_sheet_by_name("enrichment_process")
        if "collection_process" in wb.sheetnames:
            collectionSheet = wb.get_sheet_by_name("collection_process")
        if "dissociation_process" in wb.sheetnames:
            dissociationSheet = wb.get_sheet_by_name("dissociation_process")
        if "library_preparation_process" in wb.sheetnames:
            libraryPrepSheet = wb.get_sheet_by_name("library_preparation_process")
        if "sequencing_process" in wb.sheetnames:
            sequencingSheet = wb.get_sheet_by_name("sequencing_process")
        if "purchased_reagents" in wb.sheetnames:
            reagentsSheet = wb.get_sheet_by_name("purchased_reagents")
        if "sequence_file" in wb.sheetnames:
            filesSheet = wb.get_sheet_by_name("sequence_file")


        # convert data in sheets back into dict + add the schema_type/describedBy boilerplate
        project = self._multiRowToObjectFromSheet("project", projectSheet)
        if project:
            if len(project) == 1:
                project = project[0]
                project.update({"schema_type": "project",
                  "describedBy": schema_sheetname_mappings["project"]})

        enrichment = self._multiRowToObjectFromSheet("enrichment_process", enrichmentSheet)
        if enrichment:
            for e in enrichment:
                e.update({"schema_type": "process",
                          "describedBy": schema_sheetname_mappings["enrichment_process"]})

        collection = self._multiRowToObjectFromSheet("collection_process", collectionSheet)
        if collection:
            for c in collection:
                c.update({"schema_type": "process",
                     "describedBy": schema_sheetname_mappings["collection_process"]})

        dissociation = self._multiRowToObjectFromSheet("dissociation_process", dissociationSheet)
        if dissociation:
            for d in dissociation:
                d.update({"schema_type": "process",
                  "describedBy": schema_sheetname_mappings["dissociation_process"]})

        reagents = self._multiRowToObjectFromSheet("purchased_reagents", reagentsSheet)
        libraryPrep = self._multiRowToObjectFromSheet("library_preparation_process", libraryPrepSheet)
        if libraryPrep:
            for lp in libraryPrep:
                lp.update({"schema_type": "process",
                    "describedBy": schema_sheetname_mappings["library_preparation_process"]})
        sequencing = self._multiRowToObjectFromSheet("sequencing_process", sequencingSheet)
        if sequencing:
            for s in sequencing:
                s.update({"schema_type": "process",
                  "describedBy": schema_sheetname_mappings["sequencing_process"]})

        protocols = self._multiRowToObjectFromSheet("protocol", protocolSheet)
        if protocols:
            for prot in protocols:
                prot.update({"schema_type": "protocol",
                  "describedBy": schema_sheetname_mappings["protocol"]})
        donors = self._multiRowToObjectFromSheet("donor_organism", donorSheet)
        if donors:
            for do in donors:
                do.update({"schema_type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["donor_organism"]})
        familialRelationships = self._multiRowToObjectFromSheet("familial_relationship", familialRelationshipSheet)
        publications = self._multiRowToObjectFromSheet("project.publications", projectPubsSheet)
        contributors = self._multiRowToObjectFromSheet("contributor", contributorSheet)

        specimens = self._multiRowToObjectFromSheet("specimen_from_organism", specimenSheet)
        if specimens:
            for spec in specimens:
                spec.update({"schema_type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["specimen_from_organism"]})
        cell_suspension = self._multiRowToObjectFromSheet("cell_suspension", cellSuspensionSheet)
        if cell_suspension:
            for cs in cell_suspension:
                cs.update({"schema_type": "biomaterial",
                     "describedBy": schema_sheetname_mappings["cell_suspension"]})
        organoid = self._multiRowToObjectFromSheet("organoid", organoidSheet)
        if organoid:
            for org in organoid:
                org.update({"schema_type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["organoid"]})
        cell_line = self._multiRowToObjectFromSheet("cell_line", clSheet)
        if cell_line:
            for cl in cell_line:
                cl.update({"schema_type": "biomaterial",
                  "describedBy": schema_sheetname_mappings["cell_line"]})
        cell_line_publications = self._multiRowToObjectFromSheet("cell_line.publications", clPublicationSheet)
        files = self._multiRowToObjectFromSheet("sequence_file", filesSheet)
        if files:
            for f in files:
                f.update({"schema_type": "file",
                 "describedBy": schema_sheetname_mappings["sequence_file"]})


        biomaterials = []
        biomaterials.extend(donors)
        biomaterials.extend(specimens)
        biomaterials.extend(cell_suspension)
        biomaterials.extend(organoid)
        biomaterials.extend(cell_line)

        processes = []
        processes.extend(collection)
        processes.extend(dissociation)
        processes.extend(enrichment)
        processes.extend(libraryPrep)
        processes.extend(sequencing)

        chained_processes = []

        # creating submission
        #
        if not self.dryrun and not submissionUrl:
            token = "Bearer " + token_util.get_token()
            submissionUrl = self.createSubmission(token)

        # post objects to the Ingest API after some basic validation
        if existing_project_id is None:
            self.logger.info("Creating a new project for the submission")
            if "project_shortname" not in project["project_core"]:
                raise ValueError('Project must have an id attribute')
            projectId = project["project_core"]["project_shortname"]

             # embedd contact & publication into into project for now
            pubs = []
            for index, publication in enumerate(publications):
                if "project_core" in publication and "project_shortname" in publication["project_core"] and publication["project_core"]["project_shortname"] == projectId:
                    del publication["project_core"]
                else:
                    raise ValueError('Publication must reference the correct project shortname')
                pubs.append(publication)
            project["publications"] = pubs

            cont = []
            for index, contributor in enumerate(contributors):
                if "project_core" in contributor and "project_shortname" in contributor["project_core"] and contributor["project_core"]["project_shortname"] == projectId:
                    del contributor["project_core"]
                else:
                    raise ValueError('Contributor must reference the correct project shortname')
                cont.append(contributor)
            project["contributors"] = cont

            self.dumpJsonToFile(project, projectId, "project")

            projectIngest = None
            if not self.dryrun:
                projectIngest = self.ingest_api.createProject(submissionUrl, json.dumps(project), token)

        else:
            if not self.dryrun:
                self.logger.info("Retrieving existing project: " + existing_project_id)
                projectIngest = self.ingest_api.getProjectById(existing_project_id)
                submissionEnvelope = self.ingest_api.getSubmissionEnvelope(submissionUrl)
                self.ingest_api.linkEntity(projectIngest, submissionEnvelope, "submissionEnvelopes") # correct
            else:
                projectIngest = {"content" :
                                     {"project_core":
                                     {"project_shortname" : "dummy_project_id"}}}

            projectId = projectIngest["content"]["project_core"]["project_shortname"]

            self.dumpJsonToFile(projectIngest, projectId, "existing_project")

        protocolMap = {}
        for index, protocol in enumerate(protocols):
            if "protocol_id" not in protocol["protocol_core"]:
                raise ValueError('Protocol must have an id attribute')
            self.dumpJsonToFile(protocol, projectId, "protocol_" + str(index))
            protocolMap[protocol["protocol_core"]["protocol_id"]] = protocol
            if not self.dryrun:
                protocolIngest = self.ingest_api.createProtocol(submissionUrl, json.dumps(protocol))
                # self.ingest_api.linkEntity(protocolIngest, projectIngest, "projects")
                protocolMap[protocol["protocol_core"]["protocol_id"]] = protocolIngest

        biomaterialMap = {}


        for index, biomaterial in enumerate(biomaterials):
            if "biomaterial_id" not in biomaterial["biomaterial_core"]:
                raise ValueError('Biomaterial must have an id attribute')
            biomaterialMap[biomaterial["biomaterial_core"]["biomaterial_id"]] = biomaterial

            if "ncbi_taxon_id" in biomaterial["biomaterial_core"] and "genus_species" in biomaterial:
                if not isinstance(biomaterial["genus_species"],list):
                    biomaterial["genus_species"]["ontology"] = "NCBITaxon:" + str(biomaterial["biomaterial_core"]["ncbi_taxon_id"])
                elif isinstance(biomaterial["genus_species"],list) and len(biomaterial["genus_species"])==1:
                    biomaterial["genus_species"][0]["ontology"] = "NCBITaxon:" + str(biomaterial["biomaterial_core"]["ncbi_taxon_id"][0])


        # add dependent information to various biomaterial types

        for publication in cell_line_publications:
            if "biomaterial_id" in publication["biomaterial_core"]:
                bio_id = publication["biomaterial_core"]["biomaterial_id"]
                del publication["biomaterial_core"]

                if "publications" not in biomaterialMap[bio_id]:
                    biomaterialMap[bio_id]["publications"] = []
                biomaterialMap[bio_id]["publications"].append(publication)

        for familialRel in familialRelationships:
            if "biomaterial_id" in familialRel["biomaterial_core"]:
                bio_id = familialRel["biomaterial_core"]["biomaterial_id"]
                del familialRel["biomaterial_core"]

                if "familial_relationship" not in biomaterialMap[bio_id]:
                    biomaterialMap[bio_id]["familial_relationship"] = []
                biomaterialMap[bio_id]["familial_relationship"].append(familialRel)

        # build the process map from the different types of process infromation
        processMap = {}
        chainedProcessMap = {}
        for index, process in enumerate(processes):
            if "process_id" not in process["process_core"]:
                raise ValueError('Process must have an id attribute')
            processMap[process["process_core"]["process_id"]] = process

        for reagent in reagents:
            if "process_id" in reagent["process_core"]:
                process_id = reagent["process_core"]["process_id"]
                del reagent["process_core"]

                if "process_reagents" not in processMap[process_id]:
                    processMap[process_id]["process_reagents"] = []
                processMap[process_id]["process_reagents"].append(reagents)

        # submit biomaterials to ingest and link to project and protocols
        biomaterials_with_procs = []
        proc_input_biomaterials = {}
        biomaterial_proc_inputs = {}
        proc_output_biomaterials = {}
        biomaterial_proc_outputs = {}
        procs_wrapped_by = {}
        for index, biomaterial_id in enumerate(biomaterialMap.keys()):
            biomaterial = biomaterialMap[biomaterial_id]
            if "has_input_biomaterial" in biomaterial["biomaterial_core"]:
                if biomaterial["biomaterial_core"]["has_input_biomaterial"] not in biomaterialMap.keys():
                    raise ValueError('Biomaterial '+ str(biomaterial_id) +' references another biomaterial '+ str(biomaterial["biomaterial_core"]["has_input_biomaterial"]) +' that isn\'t in the spraedsheet')

            if "process_ids" in biomaterial:
                biomaterials_with_procs.append(biomaterial_id)
                # do we have multiple chained protocols? If so, create or reuse a 'wrapper'
                wrapper_process = {}
                if len(biomaterial["process_ids"]) > 1:
                    process_ids_field = str(biomaterial["process_ids"])
                    if process_ids_field in procs_wrapped_by:
                        wrapper_process = procs_wrapped_by[process_ids_field]
                    else:
                        wrapper_process = self._emptyProcessObject("wrapper", empty_wrapper_id)
                        empty_wrapper_id += 1
                        procs_wrapped_by[process_ids_field] = wrapper_process
                        processMap[wrapper_process["process_core"]["process_id"]] = wrapper_process
                        processMap[wrapper_process["process_core"]["process_id"]]["chained_process_ids"] = []

                for process_id in biomaterial["process_ids"]:
                    if process_id not in processMap.keys():
                        raise ValueError(
                         'A biomaterial references a process ' + process_id + ' that isn\'t in the biomaterials worksheet')

                    if biomaterial_id not in biomaterial_proc_inputs:
                        biomaterial_proc_inputs[biomaterial_id] = []

                    # do we have a wrapper process?
                    if wrapper_process:
                        # link this process to the wrapper
                        wrapper_id = wrapper_process["process_core"]["process_id"]
                        if "biomaterial_ids" not in processMap[wrapper_id]:
                            processMap[wrapper_id]["biomaterial_ids"] = []

                        processMap[wrapper_id]["chained_process_ids"].append(process_id)
                        if process_id not in chainedProcessMap:
                            chainedProcessMap[process_id] = []
                        if wrapper_id not in chainedProcessMap[process_id]:
                            chainedProcessMap[process_id].append(wrapper_id)

                        # link input or output biomaterials to the wrapper (i.e. indirectly)
                        if biomaterial_id not in processMap[wrapper_id]["biomaterial_ids"]:
                            processMap[wrapper_id]["biomaterial_ids"].append(biomaterial_id)
                        if wrapper_id not in proc_input_biomaterials:
                            proc_input_biomaterials[wrapper_id] = []
                        if biomaterial_id not in proc_input_biomaterials[wrapper_id]:
                            proc_input_biomaterials[wrapper_id].append(biomaterial_id)
                        if wrapper_id not in biomaterial_proc_inputs[biomaterial_id]:
                            biomaterial_proc_inputs[biomaterial_id].append(wrapper_id)
                    else:
                        # link processes to input and output directly
                        if "biomaterial_ids" not in processMap[process_id]:
                            processMap[process_id]["biomaterial_ids"] = []

                        processMap[process_id]["biomaterial_ids"].append(biomaterial["biomaterial_core"]["biomaterial_id"])
                        if process_id not in proc_input_biomaterials:
                            proc_input_biomaterials[process_id] = []
                        if biomaterial_id not in proc_input_biomaterials[process_id]:
                            proc_input_biomaterials[process_id].append(biomaterial_id)
                        if process_id not in biomaterial_proc_inputs[biomaterial_id]:
                            biomaterial_proc_inputs[biomaterial_id].append(process_id)
                del biomaterial["process_ids"]

            self.dumpJsonToFile(biomaterial, projectId, "biomaterial_" + str(index))
            if not self.dryrun:
                biomaterialIngest = self.ingest_api.createBiomaterial(submissionUrl, json.dumps(biomaterial))
                self.ingest_api.linkEntity(biomaterialIngest, projectIngest, "projects") # correct
                biomaterialMap[biomaterial["biomaterial_core"]["biomaterial_id"]] = biomaterialIngest

        # create has_input_biomaterial links between biomaterials separately to make sure all biomaterials are submitted
        for index, biomaterial_id in enumerate(biomaterialMap.keys()):
            if not self.dryrun:
                if "has_input_biomaterial" in biomaterialMap[biomaterial_id]['content']["biomaterial_core"]:
                    # retrieve biomaterials from map
                    output_biomaterial = biomaterialMap[biomaterial_id]
                    input_biomaterial = biomaterialMap[biomaterialMap[biomaterial_id]['content']["biomaterial_core"]["has_input_biomaterial"]]

                    # if the input biomaterial declares the process, we will link later
                    if input_biomaterial['content']['biomaterial_core']['biomaterial_id'] not in biomaterials_with_procs:
                        # else create sampling process to link biomaterials
                        sampling_process = self._emptyProcessObject("sampling", empty_sampling_id)
                        empty_sampling_id += 1
                        sampling_process_ingest = self.ingest_api.createProcess(submissionUrl, json.dumps(sampling_process))

                        # link process to input biomaterials
                        self.ingest_api.linkEntity(input_biomaterial, sampling_process_ingest, "inputToProcesses")
                        # link process to output biomaterials
                        self.ingest_api.linkEntity(output_biomaterial, sampling_process_ingest, "derivedByProcesses")
                    else:
                        for process_id in biomaterial_proc_inputs[input_biomaterial['content']['biomaterial_core']['biomaterial_id']]:
                            if process_id not in proc_output_biomaterials:
                                proc_output_biomaterials[process_id] = []
                            proc_output_biomaterials[process_id].append(biomaterial_id)
                            if biomaterial_id not in biomaterial_proc_outputs:
                                biomaterial_proc_outputs[biomaterial_id] = []
                            biomaterial_proc_inputs[biomaterial_id].append(process_id)

        filesMap={}
        for index, file in enumerate(files):
            if "file_name" not in file["file_core"]:
                raise ValueError('Files must have a name')
            if "process_id" not in file:
                raise ValueError('Files must be linked to a process')
            process = file["process_id"]
            del file["process_id"]
            del file["biomaterial_id"]
            filesMap[file["file_core"]["file_name"]] = file

            # is the process referred to from this file wrapped by another process?
            if process in chainedProcessMap:
                for wrapper_process in chainedProcessMap[process]:
                    if "files" not in processMap[wrapper_process]:
                        processMap[wrapper_process]["files"] = []
                    processMap[wrapper_process]["files"].append(file["file_core"]["file_name"])
            else:
                if "files" not in processMap[process]:
                    processMap[process]["files"] = []
                processMap[process]["files"].append(file["file_core"]["file_name"])

            self.dumpJsonToFile(file, projectId, "files_" + str(index))
            if not self.dryrun:
                fileIngest = self.ingest_api.createFile(submissionUrl, file["file_core"]["file_name"], json.dumps(file))
                filesMap[file["file_core"]["file_name"]] = fileIngest

        # create all the chained processes first, these will be referred to by wrapper processes
        chained_process_ingest_map = {}
        for index, chained_process in enumerate(chainedProcessMap.keys()):
            if chained_process not in processMap:
                raise ValueError('A chained process was not found in the process sheet - ' + str(chained_process))
            if not self.dryrun:
                chained_process_protocols = []
                if "protocol_ids" in processMap[chained_process]:
                    for protocol_id in processMap[chained_process]["protocol_ids"]:
                        if protocol_id not in protocolMap:
                            raise ValueError('An process references a protocol '+protocol_id+' that isn\'t in one of the protocol worksheets')
                        chained_process_protocols.append(protocolMap[protocol_id])
                    del processMap[chained_process]["protocol_ids"]

                chained_process_ingest = self.ingest_api.createProcess(submissionUrl, json.dumps(processMap[chained_process]))
                chained_process_ingest_map[chained_process] = chained_process_ingest
                for protocol in chained_process_protocols:
                    self.ingest_api.linkEntity(chained_process_ingest, protocol, "protocols")

        for index, process in enumerate(processMap.values()):
            if "process_id" not in process["process_core"]:
                raise ValueError('Each process must have an id attribute' + str(process))

            output_files=[]
            if "files" in process:
                for file in process["files"]:
                    if file not in filesMap:
                        raise ValueError('Process references file '+file+' that isn\'t defined in the files sheet')
                output_files = process["files"]
                del process["files"]

            output_biomaterials = []
            if "biomaterial_ids" not in process:
                # this is allowed if this is a chained process
                if process["process_core"]["process_id"] not in chainedProcessMap:
                    raise ValueError("Every process must reference a biomaterial using the biomaterial_id attribute")
            else:
                for biomaterial_id in process["biomaterial_ids"]:
                    if biomaterial_id not in biomaterialMap:
                        raise ValueError('An process references a biomaterial '+biomaterial_id+' that isn\'t in one of the biomaterials worksheets')
                output_biomaterials = process["biomaterial_ids"]
                del process["biomaterial_ids"]

            chained_processes=[]
            if "chained_process_ids" in process:
                for chained_process_id in process["chained_process_ids"]:
                    if not self.dryrun:
                        chained_processes.append(chained_process_ingest_map[chained_process_id])
                del process["chained_process_ids"]

            process_protocols = []
            if "protocol_ids" in process:
                for protocol_id in process["protocol_ids"]:
                    if protocol_id not in protocolMap:
                        raise ValueError('An process references a protocol '+protocol_id+' that isn\'t in one of the protocol worksheets')
                process_protocols = process["protocol_ids"]
                del process["protocol_ids"]

            self.dumpJsonToFile(process, projectId, "process_" + str(index))
            if not self.dryrun:

                process_id = process["process_core"]["process_id"]

                if process_id in chained_process_ingest_map:
                    processIngest = chained_process_ingest_map[process_id]
                else:
                    processIngest = self.ingest_api.createProcess(submissionUrl, json.dumps(process))

                self.ingest_api.linkEntity(processIngest, projectIngest, "projects") # correct

                if process["process_core"]["process_id"] in proc_input_biomaterials:
                    for biomaterial_id in proc_input_biomaterials[process["process_core"]["process_id"]]:
                        self.ingest_api.linkEntity(biomaterialMap[biomaterial_id], processIngest, "inputToProcesses")

                if process["process_core"]["process_id"] in proc_output_biomaterials:
                    for biomaterial_id in proc_output_biomaterials[process["process_core"]["process_id"]]:
                        self.ingest_api.linkEntity(biomaterialMap[biomaterial_id], processIngest, "derivedByProcesses")

                for file in output_files:
                    self.ingest_api.linkEntity(filesMap[file], processIngest, "derivedByProcesses") # correct

                for chained_process in chained_processes:
                    self.ingest_api.linkEntity(processIngest, chained_process, "chainedProcesses")  # correct

                for protocol_id in process_protocols:
                    self.ingest_api.linkEntity(processIngest, protocolMap[protocol_id], "protocols")

        self.logger.info("All done!")
        wb.close()
        return submissionUrl

if __name__ == '__main__':
    parser = OptionParser()
    parser.add_option("-p", "--path", dest="path",
                      help="path to HCA example data bundles", metavar="FILE")
    parser.add_option("-d", "--dry", help="doa dry run without submitting to ingest", action="store_true", default=False)
    parser.add_option("-o", "--output", dest="output",
                      help="output directory where to dump json files submitted to ingest", metavar="FILE", default=None)
    parser.add_option("-i", "--id", dest="project_id",
                      help="The project_id for an existing submission", default=None)
    parser.add_option("-v", "--version", dest="schema_version", help="Metadata schema version", default=None)


    (options, args) = parser.parse_args()
    if not options.path:
        print ("You must supply path to the HCA bundles directory")
        exit(2)
    submission = SpreadsheetSubmission(options.dry, options.output, options.schema_version)
    submission.submit(options.path, None, None, options.project_id)
