import os
from unittest.case import TestCase

from mock import MagicMock, patch, Mock
from openpyxl import Workbook

from ingest.api.ingestapi import IngestApi
from ingest.importer.conversion.metadata_entity import MetadataEntity
from ingest.importer.importer import WorksheetImporter, WorkbookImporter, MultipleProjectsFound, \
    NoProjectFound, SchemaRetrievalError, WorkbookImportingException
from ingest.importer.importer import XlsImporter
from ingest.importer.spreadsheet.ingest_workbook import IngestWorkbook, IngestWorksheet
from ingest.utils.IngestError import ImporterError
from tests.importer.utils.test_utils import create_test_workbook

BASE_PATH = os.path.dirname(__file__)

HEADER_ROW = 4


def _create_single_row_worksheet(worksheet_data: dict):
    workbook = Workbook()
    worksheet = workbook.create_sheet()

    for column, data in worksheet_data.items():
        key, value = data
        worksheet[f'{column}{HEADER_ROW}'] = key
        worksheet[f'{column}6'] = value

    return worksheet


class XlsImporterTest(TestCase):

    @patch('ingest.importer.importer.IngestWorkbook')
    @patch('ingest.importer.importer.WorkbookImporter')
    @patch('ingest.importer.importer.template_manager')
    def test_generate_json_success(self, mock_template_manager, mock_wb_importer, mock_wb):
        mock_template_manager.build = MagicMock('template_manager', return_value='template_manager')
        mock_wb_importer.return_value.do_import = Mock(return_value=({'test': 'output'}, ['errors']))
        ingest_api = MagicMock('ingest_api')
        importer = XlsImporter(ingest_api)
        spreadsheet_json, template_manager, errors = importer.generate_json('file_path')

        self.assertEqual(spreadsheet_json, {'test': 'output'}, )
        self.assertEqual(errors, ['errors'])
        self.assertEqual(template_manager, 'template_manager')

    @patch('ingest.importer.importer.IngestWorkbook')
    @patch('ingest.importer.importer.WorkbookImporter')
    @patch('ingest.importer.importer.template_manager')
    def test_generate_json_error(self, mock_template_manager, mock_wb_importer, mock_wb):
        mock_template_manager.build = MagicMock('template_manager', side_effect=Exception('test error'))
        ingest_api = MagicMock('ingest_api')
        importer = XlsImporter(ingest_api)
        with self.assertRaises(SchemaRetrievalError):
            importer.generate_json('file_path')

    @patch('ingest.importer.submission.EntityMap.load')
    @patch('ingest.importer.submission.EntityLinker.process_links_from_spreadsheet')
    @patch('ingest.importer.importer.XlsImporter.generate_json')
    def test_dry_run_import_file_success(self, mock_generate_json, mock_entity_linker, mock_entity_map):
        mock_entity_map.return_value = MagicMock()
        mock_entity_linker.return_value = 'entity_map_w_links'
        mock_generate_json.return_value = ({'test': 'output'}, 'template_manager', [])
        ingest_api = MagicMock('ingest_api')
        importer = XlsImporter(ingest_api)

        entity_map = importer.dry_run_import_file('file_path')
        self.assertEqual(entity_map, 'entity_map_w_links')

    @patch('ingest.importer.submission.EntityMap.load')
    @patch('ingest.importer.submission.EntityLinker.process_links_from_spreadsheet')
    @patch('ingest.importer.importer.XlsImporter.generate_json')
    def test_dry_run_import_file_error(self, mock_generate_json, mock_entity_linker, mock_entity_map):
        mock_entity_map.return_value = MagicMock()
        mock_entity_linker.return_value = 'entity_map_w_links'
        mock_generate_json.return_value = ({'test': 'output'}, 'template_manager', ['error'])
        ingest_api = MagicMock('ingest_api')
        importer = XlsImporter(ingest_api)
        with self.assertRaises(WorkbookImportingException) as e:
            importer.dry_run_import_file('file_path')
            self.assertEqual(e.errors, ['error'])


class WorkbookImporterTest(TestCase):
    mock_json_schemas = [
        {'name': 'project', 'properties': ['contributors']},
        {'name': 'users', 'properties': ['sn_profiles']}
    ]

    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        template_mgr.template.json_schemas = self.mock_json_schemas
        template_mgr.get_concrete_type = MagicMock(side_effect=['project', 'users'])

        worksheet_importer = WorksheetImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer
        no_errors = []
        # and:
        project = MetadataEntity(concrete_type='project', domain_type='project')
        jdelacruz = MetadataEntity(concrete_type='user', domain_type='user', object_id=1,
                                   content={'user_name': 'jdelacruz'})
        setsuna_f_seiei = MetadataEntity(concrete_type='user', domain_type='user', object_id=96,
                                         content={'user_name': 'sayyeah'})
        worksheet_importer.do_import = MagicMock(
            side_effect=[([project], no_errors), ([jdelacruz, setsuna_f_seiei], no_errors)])

        # and:
        workbook = create_test_workbook('Project', 'Users')
        ingest_workbook = IngestWorkbook(workbook)
        workbook_importer = WorkbookImporter(template_mgr)

        # when:
        workbook_json, errors = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertIsNotNone(workbook_json)
        self.assertEqual(errors, [])

        # and:
        user_map = workbook_json.get('user')
        self.assertIsNotNone(user_map)
        self.assertEqual(2, len(user_map))
        self.assertEqual([jdelacruz.object_id, setsuna_f_seiei.object_id], list(user_map.keys()))

        # and:
        self.assertEqual({'user_name': 'jdelacruz'}, user_map.get(1)['content'])
        self.assertEqual({'user_name': 'sayyeah'}, user_map.get(96)['content'])

    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import_with_module_tab(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        template_mgr.template.json_schemas = self.mock_json_schemas
        template_mgr.get_concrete_type = MagicMock(side_effect=['project', 'users', 'users'])

        worksheet_importer = WorksheetImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer
        no_errors = []
        # and: stub worksheet importer
        project = MetadataEntity(concrete_type='Project', domain_type='Project')
        user = MetadataEntity(concrete_type='user', domain_type='user', object_id=773,
                              content={'user_name': 'janedoe'})
        fb_profile = MetadataEntity(concrete_type='sn_profile', domain_type='user', object_id=773,
                                    content={'sn_profiles': {'name': 'facebook', 'id': '392'},
                                             'description': 'extra fb field'})
        ig_profile = MetadataEntity(concrete_type='sn_profile', domain_type='user', object_id=773,
                                    content={'sn_profiles': {'name': 'instagram', 'id': 'a92'},
                                             'description': 'extra ig field'})
        expected_errors = [
            {'key': 'description', 'value': 'extra fb field'},
            {'key': 'description', 'value': 'extra ig field'}
        ]

        worksheet_importer.do_import = MagicMock(side_effect=[
            ([project], no_errors),
            ([user], no_errors),
            ([fb_profile, ig_profile], no_errors)])

        # and: create test workbook
        workbook = create_test_workbook('Project', 'User', 'User - SN Profiles')
        ingest_workbook = IngestWorkbook(workbook)
        workbook_importer = WorkbookImporter(template_mgr)

        # when:
        spreadsheet_json, errors = workbook_importer.do_import(ingest_workbook)

        # then:
        self.assertIsNotNone(spreadsheet_json)
        self.assertEqual(errors, workbook_importer.list_data_removal_errors('User - SN Profiles', expected_errors))
        self.assertEqual(2, len(spreadsheet_json))

        # and:
        user_map = spreadsheet_json.get('user')
        self.assertIsNotNone(user_map)

        # and:
        janedoe = user_map.get(773)
        self.assertIsNotNone(janedoe)
        content = janedoe.get('content')
        self.assertEqual('janedoe', content.get('user_name'))
        self.assertEqual(['user_name', 'sn_profiles'], list(content.keys()))

        # and:
        sn_profiles = content.get('sn_profiles')
        self.assertIsNotNone(sn_profiles)
        self.assertEqual(2, len(sn_profiles))

        # and:
        self.assertEqual({'name': 'facebook', 'id': '392'}, sn_profiles[0])
        self.assertEqual({'name': 'instagram', 'id': 'a92'}, sn_profiles[1])

    # NOTE: this is added because the team chose not to define an identity label for Project schema
    # This breaks the schema agnosticism of the importer framework.
    # The assumption is that there's only one Project per submission
    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import_project_worksheet(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        template_mgr.template.json_schemas = self.mock_json_schemas
        template_mgr.get_concrete_type = MagicMock(return_value='project')

        worksheet_importer = WorkbookImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer
        no_errors = []

        # and:
        project = MetadataEntity(domain_type='project', concrete_type='project',
                                 content={'description': 'test project'})
        jsmith = MetadataEntity(domain_type='project', concrete_type='contact',
                                content={'contributors': {'name': 'John',
                                                          'email': 'jsmith@email.com'}})
        ppan = MetadataEntity(domain_type='project', concrete_type='contact',
                              content={'contributors': {'name': 'Peter',
                                                        'email': 'peterpan@email.com'}})
        worksheet_importer.do_import = MagicMock(side_effect=[([project], no_errors), ([jsmith, ppan], no_errors)])

        # and:
        workbook = create_test_workbook('Project', 'Project - Contributors')
        ingest_workbook = IngestWorkbook(workbook)
        workbook_importer = WorkbookImporter(template_mgr)

        # when:
        spreadsheet_json, errors = workbook_importer.do_import(ingest_workbook)

        # then:
        project_map = spreadsheet_json.get('project')
        self.assertEqual(1, len(project_map))
        project_content = list(project_map.values())[0].get('content')
        self.assertEqual('test project', project_content.get('description'))
        self.assertEqual(errors, [])

        # and:
        contributors = project_content.get('contributors')
        self.assertEqual(2, len(contributors))
        self.assertIn({'name': 'John', 'email': 'jsmith@email.com'}, contributors)
        self.assertIn({'name': 'Peter', 'email': 'peterpan@email.com'}, contributors)

    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import_multiple_projects(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        template_mgr.template.json_schemas = self.mock_json_schemas
        template_mgr.get_concrete_type = MagicMock(return_value='project')

        worksheet_importer = WorksheetImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer
        no_errors = []
        expected_error = {
            'location': 'sheet=Project',
            'type': 'MultipleProjectsFound',
            'detail': 'The spreadsheet should only be associated to a single project.'
        }

        # and:
        project_1 = MetadataEntity(concrete_type='project', domain_type='project', object_id=1)
        project_2 = MetadataEntity(concrete_type='project', domain_type='project', object_id=2)
        worksheet_importer.do_import = MagicMock(side_effect=[([project_1, project_2], no_errors)])

        # and:
        workbook = create_test_workbook('Project')
        workbook_importer = WorkbookImporter(template_mgr)

        # when:
        spreadsheet_json, errors = workbook_importer.do_import(IngestWorkbook(workbook))

        # then:
        self.assertIn(expected_error, errors, f'Errors expected to contain {MultipleProjectsFound.__name__}.')

    @patch('ingest.importer.importer.WorksheetImporter')
    def test_do_import_no_projects(self, worksheet_importer_constructor):
        # given:
        template_mgr = MagicMock(name='template_manager')
        worksheet_importer = WorksheetImporter(template_mgr)
        worksheet_importer_constructor.return_value = worksheet_importer
        no_errors = []
        expected_error = {
            'location': 'File',
            'type': 'NoProjectFound',
            'detail': 'The spreadsheet should be associated to a project.'
        }

        # and:
        item = MetadataEntity(concrete_type='product', domain_type='product', object_id=910)
        worksheet_importer.do_import = MagicMock(side_effect=[([item], no_errors)])

        # and:
        workbook = create_test_workbook('Item')
        workbook_importer = WorkbookImporter(template_mgr)

        # when:
        spreadsheet_json, errors = workbook_importer.do_import(IngestWorkbook(workbook))

        # then:
        self.assertIn(expected_error, errors, f'Errors expected to contain {NoProjectFound.__name__}.')


class WorksheetImporterTest(TestCase):

    def test_do_import(self):
        # given:
        row_template = MagicMock('row_template')
        no_errors = []

        # and:
        john_doe = MetadataEntity(object_id='profile_1')
        emma_jackson = MetadataEntity(object_id='profile_2')
        row_template.do_import = MagicMock('import_row', side_effect=[(john_doe, no_errors), (emma_jackson, no_errors)])

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('user_profile')
        worksheet['A4'] = 'header'
        worksheet['A6'] = 'john'
        worksheet['A7'] = 'emma'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        profiles, errors = worksheet_importer.do_import(IngestWorksheet(worksheet))

        # then:
        self.assertEqual(2, len(profiles))
        self.assertIn(john_doe, profiles)
        self.assertIn(emma_jackson, profiles)
        self.assertEqual(errors, [])

        # and: domain and concrete type should be set
        pass

    def test_do_import_no_id_metadata(self):
        # given:
        row_template = MagicMock('row_template')
        no_errors = []
        # and:
        paper_metadata = MetadataEntity(content={'product_name': 'paper'},
                                        links={'delivery': ['123', '456']})
        pen_metadata = MetadataEntity(content={'product_name': 'pen'},
                                      links={'delivery': ['789']})
        row_template.do_import = MagicMock(side_effect=([(paper_metadata, no_errors), (pen_metadata, no_errors)]))

        # and:
        mock_template_manager = MagicMock('template_manager')
        mock_template_manager.create_row_template = MagicMock(return_value=row_template)
        mock_template_manager.get_header_row = MagicMock(return_value=['header1', 'header2'])
        mock_template_manager.get_concrete_type = MagicMock(return_value='concrete_entity')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('product')
        worksheet['A6'] = 'paper'
        worksheet['A7'] = 'pen'

        # when:
        worksheet_importer = WorksheetImporter(mock_template_manager)
        results, errors = worksheet_importer.do_import(IngestWorksheet(worksheet))

        # then:
        self.assertEqual(2, len(results))
        self.assertIn(paper_metadata, results)
        self.assertIn(pen_metadata, results)
        self.assertEqual(errors, [])

        # and: object id should be assigned
        paper_id = paper_metadata.object_id
        self.assertIsNotNone(paper_id)
        pen_id = pen_metadata.object_id
        self.assertIsNotNone(pen_id)
        self.assertNotEqual(paper_id, pen_id)


class ImporterErrors(TestCase):
    def setUp(self):
        # Setup mocked APIs
        self.mock_ingest_api = MagicMock(spec=IngestApi)

    def test_import_file_throwing_exception(self):
        # given:
        exception = Exception('Error thrown for Unit Test')
        exception_json = ImporterError(str(exception)).getJSON()
        importer = XlsImporter(ingest_api=self.mock_ingest_api)
        importer.generate_json = MagicMock(side_effect=exception)
        importer.logger.error = MagicMock()

        # when:
        importer.import_file(file_path=None, submission_url=None)
        # then:
        self.mock_ingest_api.create_submission_error.assert_called_once_with(None, exception_json)
