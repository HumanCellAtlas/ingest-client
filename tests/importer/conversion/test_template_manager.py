from unittest import TestCase

from mock import MagicMock, patch
from openpyxl import Workbook

from ingest.importer.conversion import conversion_strategy
from ingest.importer.conversion.conversion_strategy import CellConversion
from ingest.importer.conversion.template_manager import TemplateManager, RowTemplate, InvalidTabName
from ingest.importer.data_node import DataNode
from ingest.importer.spreadsheet.ingest_worksheet import IngestWorksheet, \
    IngestRow
from tests.importer.utils.test_utils import create_test_workbook


def _mock_schema_template_lookup(value_type='string', multivalue=False):
    schema_template = MagicMock(name='schema_template')
    single_string_spec = {
        'value_type': value_type,
        'multivalue': multivalue
    }
    schema_template.lookup = MagicMock(name='lookup', return_value=single_string_spec)
    return schema_template


class TemplateManagerTest(TestCase):

    def test_create_template_node(self):
        # given:
        schema_template = MagicMock(name='schema_template')
        schema_template.lookup_metadata_schema_name_given_title = MagicMock(
            return_value='concrete_entity')

        schema_url = 'https://schema.humancellatlas.org/type/biomaterial/5.1.0/donor_organsim'

        lookup_map = {
            'concrete_entity': {
                'schema': {
                    'domain_entity': 'biomaterial',
                    'url': schema_url
                }
            }
        }

        schema_template.lookup_property_from_template = lambda key: lookup_map.get(key)

        ingest_api = MagicMock(name='mock_ingest_api')

        # and:
        template_manager = TemplateManager(schema_template, ingest_api)

        # and:
        workbook = Workbook()
        donor_worksheet = workbook.create_sheet('Donor')

        # when:
        data_node: DataNode = template_manager.create_template_node(donor_worksheet)

        # then:
        data = data_node.as_dict()
        self.assertEqual(schema_url, data.get('describedBy'))
        self.assertEqual('biomaterial', data.get('schema_type'))

    # TODO move the logic of creating the column spec to SchemaTemplate
    @patch.object(conversion_strategy, 'determine_strategy')
    def test_create_row_template(self, determine_strategy):
        # given:
        template = MagicMock(name='schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        # and:
        concrete_type = "user"
        domain_entity = "main_category/subdomain"
        schema_url = "http://schema.sample.com/main_category"
        template.get_tabs_config = MagicMock()
        template.lookup_metadata_schema_name_given_title = MagicMock(return_value=concrete_type)
        template.get_latest_schema = MagicMock(return_value=schema_url)
        schema = {
            "schema": {
                "domain_entity": domain_entity,
                "url": schema_url
            }
        }
        property_one_schema = {
            "description": "Property one description",
            "value_type": "string"
        }
        property_two_schema = {
            "description": "Property two description",
            "value_type": "string"
        }
        spec_map = {
            concrete_type: schema,
            "user.profile.first_name": property_one_schema,
            "user.numbers": property_two_schema

        }
        template.lookup_property_from_template = lambda key: spec_map.get(key)

        # and:
        name_strategy = MagicMock('name_strategy')
        numbers_strategy = MagicMock('numbers_strategy')
        determine_strategy.side_effect = [name_strategy, numbers_strategy]

        # and: prepare worksheet
        header_row_idx = 4
        workbook = Workbook()
        worksheet = workbook.create_sheet('sample')
        worksheet[f'A{header_row_idx}'] = 'user.profile.first_name'
        worksheet[f'B{header_row_idx}'] = 'user.numbers'

        ingest_worksheet = IngestWorksheet(worksheet, header_row_idx=header_row_idx)

        # when:
        template_manager = TemplateManager(template, ingest_api)
        row_template: RowTemplate = template_manager.create_row_template(ingest_worksheet)

        # then:
        self.assertEqual(determine_strategy.call_count, 2)

        # and:
        self.assertIsNotNone(row_template)
        self.assertEqual('main_category', row_template.domain_type)
        self.assertEqual(concrete_type, row_template.concrete_type)
        self.assertEqual(2, len(row_template.cell_conversions))
        self.assertTrue(name_strategy in row_template.cell_conversions)
        self.assertTrue(numbers_strategy in row_template.cell_conversions)

    @patch.object(conversion_strategy, 'determine_strategy')
    def test_create_row_template_with_default_values(self, determine_strategy):
        # given:
        schema_template = MagicMock('schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        # and:
        domain_entity = "profile/profile_type"
        schema_url = "http://schema.sample.com/profile"
        schema_template.get_tabs_config = MagicMock()
        schema_template.lookup_metadata_schema_name_given_title = MagicMock(
            return_value="profile_type")
        schema_template.get_latest_schema = MagicMock(return_value=schema_url)
        schema = {
            "schema": {
                "domain_entity": domain_entity,
                "url": schema_url
            }
        }
        property_schema = {
            "description": "Property description",
            "value_type": "string"
        }
        spec_map = {
            "profile_type": schema,
            "profile.name": property_schema
        }
        schema_template.lookup_property_from_template = lambda key: spec_map.get(key)

        # and:
        determine_strategy.return_value = FakeConversion('')

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('profile')
        worksheet['A4'] = 'profile.name'
        ingest_worksheet = IngestWorksheet(worksheet)

        # when:
        template_manager = TemplateManager(schema_template, ingest_api)
        template_manager.get_schema_url = MagicMock(return_value=schema_url)
        row_template = template_manager.create_row_template(ingest_worksheet)

        # then:
        content_defaults = row_template.default_values
        self.assertIsNotNone(content_defaults)
        self.assertEqual(schema_url, content_defaults.get('describedBy'))
        self.assertEqual('profile', content_defaults.get('schema_type'))

    @patch.object(conversion_strategy, 'determine_strategy')
    def test_create_row_template_for_module_worksheet(self, determine_strategy):
        # given:
        template = MagicMock(name='schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        # and:
        spec_map = {
            'product': {'schema': {'domain_entity': 'merchandise/product'}}
        }
        template.lookup_property_from_template = lambda key: spec_map.get(key, None)

        domain_entity = "merchandise/product"
        schema_url = "http://schema.sample.com/product"
        template.get_tabs_config = MagicMock()
        template.lookup_metadata_schema_name_given_title = MagicMock(
            return_value="product_type")
        template.get_latest_schema = MagicMock(return_value=schema_url)
        schema = {
            "schema": {
                "domain_entity": domain_entity,
                "url": schema_url
            }
        }
        property_one_schema = {
            "description": "Property one description",
            "value_type": "string"
        }
        property_two_schema = {
            "description": "Property two description",
            "value_type": "string"
        }
        spec_map = {
            "product_type": schema,
            "product.info.id": property_one_schema,
            "product.reviews.rating": property_two_schema
        }
        template.lookup_property_attributes_in_metadata = lambda key: spec_map.get(key)

        # and:
        template_mgr = TemplateManager(template, ingest_api)

        # and:
        workbook = create_test_workbook('Product - Reviews')
        reviews_worksheet = workbook['Product - Reviews']
        reviews_worksheet['A4'] = 'product.info.id'
        reviews_worksheet['B4'] = 'product.reviews.rating'

        # and: set up strategies
        id_strategy = MagicMock(name='id_strategy')
        rating_strategy = MagicMock(name='rating_strategy')
        determine_strategy.side_effect = [id_strategy, rating_strategy]

        # when:
        row_template = template_mgr.create_row_template(IngestWorksheet(reviews_worksheet))

        # and:
        self.assertIsNotNone(row_template)
        self.assertIn(id_strategy, row_template.cell_conversions)
        self.assertIn(rating_strategy, row_template.cell_conversions)

    @patch.object(conversion_strategy, 'determine_strategy')
    def test_create_row_template_with_none_header(self, determine_strategy):
        # given:
        schema_template = MagicMock('schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        # and:
        do_nothing_strategy = FakeConversion('')
        determine_strategy.return_value = do_nothing_strategy

        # and:
        self._mock_schema_lookup(schema_template)

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('sample')
        worksheet['A4'] = None
        ingest_worksheet = IngestWorksheet(worksheet)

        # when:
        template_manager = TemplateManager(schema_template, ingest_api)
        row_template = template_manager.create_row_template(ingest_worksheet)

        # then:
        self.assertEqual(0, len(row_template.cell_conversions))

    @staticmethod
    def _mock_schema_lookup(schema_template, schema_url='', object_type='', main_category=None):
        schema_template.get_tabs_config = MagicMock()
        schema_template.lookup_metadata_schema_name_given_title = MagicMock(
            return_value=object_type)
        schema_template.get_latest_schema = MagicMock(return_value=schema_url)

        domain_entity = f'{main_category}/{object_type}' if main_category else object_type
        schema = {'schema': {
            'domain_entity': domain_entity,
            'url': schema_url
        }}
        spec_map = {
            object_type: schema
        }
        schema_template.lookup_property_from_template = lambda key: spec_map.get(key)

    def test_get_schema_type(self):
        # given
        schema_template = MagicMock(name='schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        spec = {
            'schema': {
                'high_level_entity': 'type',
                'domain_entity': 'biomaterial',
                'module': 'donor_organism',
                'url': 'https://schema.humancellatlas.org/type/biomaterial/5.0.0/donor_organism'
            }
        }
        schema_template.lookup_property_from_template = MagicMock(
            name='lookup_property_from_template', return_value=spec)
        template_manager = TemplateManager(schema_template, ingest_api)

        # when:
        domain_entity = template_manager.get_domain_type('cell_suspension')

        # then:
        self.assertEqual('biomaterial', domain_entity)

    def test_get_schema_url(self):
        # given
        latest_url = 'https://schema.humancellatlas.org/type/biomaterial/5.0.0/donor_organism'
        schema_template = MagicMock(name='schema_template')
        ingest_api = MagicMock(name='mock_ingest_api')

        spec = {
            'schema': {
                'high_level_entity': 'type',
                'domain_entity': 'biomaterial',
                'module': 'donor_organism',
                'url': latest_url
            }
        }

        schema_template.lookup_property_from_template = MagicMock(
            name='lookup_property_from_template', return_value=spec)
        template_manager = TemplateManager(schema_template, ingest_api)
        template_manager.get_latest_schema_url = MagicMock(return_value=latest_url)

        # when:
        url = template_manager.get_schema_url('cell_suspension')

        # then:
        self.assertEqual('https://schema.humancellatlas.org/type/biomaterial/5.0.0/donor_organism', url)

    def test_get_concrete_type_of_regular_worksheet(self):
        # given
        schema_template = MagicMock(name='schema_template')
        schema_template.lookup_metadata_schema_name_given_title = MagicMock(
            return_value='user_profile')
        manager = TemplateManager(schema_template, MagicMock(name='mock_ingest_api'))

        # expect:
        self.assertEqual('user_profile', manager.get_concrete_type('User Profile'))
        schema_template.lookup_metadata_schema_name_given_title.assert_called_with('User Profile')

    def test_get_concrete_type_of_module_worksheet(self):
        # given:
        schema_template = MagicMock(name='schema_template')
        schema_template.lookup_metadata_schema_name_given_title = MagicMock(return_value='product')
        manager = TemplateManager(schema_template, MagicMock(name='mock_ingest_api'))

        # expect:
        self.assertEqual('product', manager.get_concrete_type('Product - Barcodes'))

        # and:
        schema_template.lookup_metadata_schema_name_given_title.assert_called_with('Product')

    def test_get_concrete_type_of_worksheet_invalid_format(self):
        # given:
        schema_template = MagicMock(name='schema_template')
        manager = TemplateManager(schema_template, MagicMock(name='mock_ingest_api'))

        # when:
        raised_exception = None
        try:
            manager.get_concrete_type('- does not match format -')
        except InvalidTabName as exception:
            raised_exception = exception

        # then:
        self.assertIsNotNone(raised_exception)
        self.assertEqual('- does not match format -', raised_exception.tab_name)

    def test_get_domain_type(self):
        # given:
        template = MagicMock(name='schema_template')
        schema_spec = {'schema': {'domain_entity': 'user/profile'}}
        template.lookup_property_from_template = MagicMock(return_value=schema_spec)

        # and:
        template_manager = TemplateManager(template, MagicMock(name='mock_ingest_api'))

        # expect:
        self.assertEqual('user', template_manager.get_domain_type('profile'))


class FakeConversion(CellConversion):
    def __init__(self, field):
        self.field = field

    def apply(self, metadata, cell_data):
        metadata.define_content(self.field, cell_data)


class BrokenConversion(CellConversion):
    def __init__(self, field):
        self.field = field

    def apply(self, metadata, cell_data):
        raise Exception('Unit Test Error converting cell data.')


class RowTemplateTest(TestCase):
    def test_do_import(self):
        # given:
        cell_conversions = [FakeConversion('first_name'), FakeConversion('last_name'),
                            FakeConversion('address.city'), FakeConversion('address.country')]
        row_template = RowTemplate('user', 'user_profile', cell_conversions)

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('profile')
        worksheet['A1'] = 'Juan'
        worksheet['B1'] = 'dela Cruz'
        worksheet['C1'] = 'Manila'
        worksheet['D1'] = 'Philippines'
        row = list(worksheet.rows)[0]
        row = IngestRow('worksheet_title', 0, row)

        # when:
        result, errors = row_template.do_import(row)

        # then:
        self.assertIsNotNone(result)
        self.assertEqual('user', result.domain_type)
        self.assertEqual('user_profile', result.concrete_type)
        self.assertEqual('Juan', result.get_content('first_name'))
        self.assertEqual('dela Cruz', result.get_content('last_name'))

        # and:
        address = result.get_content('address')
        self.assertIsNotNone(address)
        self.assertEqual('Manila', address.get('city'))
        self.assertEqual('Philippines', address.get('country'))
        self.assertEqual(errors, [])

    def test_do_import_with_default_values(self):
        # given:
        schema_url = 'https://sample.com/list_of_things'
        default_values = {
            'describedBy': schema_url,
            'extra_field': 'extra field'
        }

        # and:
        cell_conversions = [FakeConversion('name'), FakeConversion('description')]
        row_template = RowTemplate('user', 'user', cell_conversions, default_values=default_values)

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('list_of_things')
        worksheet['A1'] = 'pen'
        worksheet['B1'] = 'a thing used for writing'
        row = list(worksheet.rows)[0]
        row = IngestRow('worksheet_title', 0, row)

        # when:
        result, errors = row_template.do_import(row)

        # then:
        self.assertEqual(schema_url, result.get_content('describedBy'))
        self.assertEqual('extra field', result.get_content('extra_field'))
        self.assertEqual(errors, [])

    def test_do_import_with_conversion_error(self):
        # given:
        cell_conversions = [FakeConversion('first_name'), FakeConversion('last_name'),
                            BrokenConversion('address.city'), FakeConversion('address.country')]
        row_template = RowTemplate('user', 'user_profile', cell_conversions)
        expected_error = {
            'location': 'column=2, value=Manila',
            'type': 'Exception',
            'detail': 'Unit Test Error converting cell data.'
        }

        # and:
        workbook = Workbook()
        worksheet = workbook.create_sheet('profile')
        worksheet['A1'] = 'Juan'
        worksheet['B1'] = 'dela Cruz'
        worksheet['C1'] = 'Manila'
        worksheet['D1'] = 'Philippines'
        row = list(worksheet.rows)[0]
        row = IngestRow('worksheet_title', 0, row)

        # when:
        result, errors = row_template.do_import(row)

        # then:
        self.assertIn(expected_error, errors)
        self.assertIsNotNone(result)
        self.assertEqual('user', result.domain_type)
        self.assertEqual('user_profile', result.concrete_type)
        self.assertEqual('Juan', result.get_content('first_name'))
        self.assertEqual('dela Cruz', result.get_content('last_name'))

        # and:
        address = result.get_content('address')
        self.assertIsNotNone(address)
        self.assertEqual('Philippines', address.get('country'))
