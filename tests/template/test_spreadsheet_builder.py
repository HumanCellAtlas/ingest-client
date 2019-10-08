import os

from ingest.template.schema_template import SchemaTemplate

__author__ = "jupp"
__license__ = "Apache 2.0"
__date__ = "25/05/2018"

from unittest import TestCase
from ingest.template.vanilla_spreadsheet_builder import VanillaSpreadsheetBuilder
import unittest
from openpyxl import load_workbook as Reader


class TestSchemaTemplate(TestCase):

    def setUp(self):
        self.longMessage = True
        self.projectUri = "https://schema.humancellatlas.org/type/project/5.1.0/project"
        self.donorUri = "https://schema.humancellatlas.org/type/biomaterial/5.1.0/donor_organism"

    def test_no_schemas(self):
        data = {
            "id": self.donorUri,
            "properties": {
                "foo_bar": {
                    "user_friendly": "Foo bar",
                    "description": "this is a foo bar",
                    "example": "e.g. foo"
                }
            }
        }

        file = "foo.xlsx"
        spreadsheet_builder = VanillaSpreadsheetBuilder(file)
        template = SchemaTemplate(json_schema_docs=[data], property_migrations=[])
        spreadsheet_builder.build(template)
        spreadsheet_builder.save_spreadsheet()

        reader = Reader("foo.xlsx")
        sheet = reader["Donor organism"]

        self.assertEqual("this is a foo bar", sheet.cell(row=2, column=1).value)
        self.assertEqual("FOO BAR", sheet.cell(row=1, column=1).value)
        self.assertEqual("For example: e.g. foo", sheet.cell(row=3, column=1).value.strip())
        self.assertEqual("donor_organism.foo_bar", sheet.cell(row=4, column=1).value)
        # clean up
        os.remove(file)

    # TODO fixme
    @unittest.skip
    def test_with_tabs_template(self):
        pass
        # spreadsheet_builder.generate_spreadsheet("human_10x.xlsx", tabs_template="tabs_human_10x.yaml",
        #                                          schema_urls=schemas)

    # TODO fixme
    @unittest.skip
    def test_add_columns(self):
        # spreadsheet_builder.generate_spreadsheet("generic.xlsx", schema_urls=schemas)
        pass

    # TODO fixme
    @unittest.skip
    def test_add_sheets(self):
        # spreadsheet_builder.generate_spreadsheet("generic.xlsx", schema_urls=schemas)
        pass
