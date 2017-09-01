import glob, json, os, urllib, requests
from openpyxl import load_workbook
from ingestapi import IngestApi
from optparse import OptionParser


# This script will read a spreadsheet, generate a manifest, submit all items to the ingest API, assign uuid and generate a directory of bundles for the
# submitted data

parser = OptionParser()
parser.add_option("-p", "--path", dest="path",
                  help="path to HCA fromatted spreadsheet", metavar="FILE")

parser.add_option("-u", "--url",
                  help="base URL to the ingest API")

(options, args) = parser.parse_args()
if not options.path:
    print "You must supply an excel file"
    exit(2)

class MetadataDocument:
    def __init__(self, content):
        self.content = content

    def toJSON(self):
        return json.dumps(self, default=lambda o: o.__dict__,
            sort_keys=True, indent=4)

wb = load_workbook(filename = options.path)

projectSheet = wb.get_sheet_by_name("project")
contactSheet = wb.get_sheet_by_name("contact")
sampleSheet = wb.get_sheet_by_name("sample")
donorSheet = wb.get_sheet_by_name("donor")
protocolSheet = wb.get_sheet_by_name("protocols")
assaySheet = wb.get_sheet_by_name("assay")

def keyValueToNestedObject(key, value):
    d = value
    if "\"" in unicode(value) or "||" in unicode(value):
        d = map(lambda it: it.strip("\""), value.split("||"))

    for part in reversed(key.split('.')):
        d = {part: d}
    return d

def multiRowToObjectFromSheet(type, sheet):
    objs = []
    for row in sheet.iter_rows(row_offset=1, max_row=(sheet.max_row -1)):

        obj = {}
        for cell in row:
            if not cell.value:
                continue
            cellCol = cell.col_idx
            propertyValue = sheet.cell(row=1, column=cellCol).value
            d = keyValueToNestedObject(propertyValue, cell.value)
            obj.update(d)
        print json.dumps(obj)
        objs.append(obj)

    return objs

# sheets that represent one entity where the properties are in column 0
def sheetToObject(type, sheet):
    obj = {}
    for row in sheet.iter_rows():
        propertyCell = row[0].value
        valueCell = row[1].value
        if valueCell:
            d = keyValueToNestedObject(propertyCell, valueCell)
            obj.update(d)
    print json.dumps(obj)
    return obj



project = sheetToObject("project", projectSheet)
contact = sheetToObject("contact", contactSheet)
project["contact"] = contact

samples = multiRowToObjectFromSheet("sample",sampleSheet)
protocols = multiRowToObjectFromSheet("protocol",protocolSheet)

donors = multiRowToObjectFromSheet("donor",donorSheet)

assays = multiRowToObjectFromSheet("assay",assaySheet)

projectId = project["id"]

dir = "./pre-ingest-example/"+projectId
if not os.path.exists(dir):
    os.makedirs(dir)

def dumpJsonToFile(object, dir, name):
    tmpFile = open(dir+"/"+name+".json", "w")
    tmpFile.write(object)
    tmpFile.close()


ingest_api = IngestApi()
print "creating submission..."

ingest_api.createSubmission()
print "new submission " + ingest_api.currentSubmission

# creating submission

# projectObj = MetadataDocument(content=project)
dumpJsonToFile(json.dumps(project) , dir, "project" )
ingest_api.createProject(json.dumps(project))

for index, sample in enumerate(samples):

    sample["protocols"] = protocols
    # sampleObj = MetadataDocument(sample)
    dumpJsonToFile(json.dumps(sample), dir, "sample_"+str(index))
    ingest_api.createSample(json.dumps(sample))

for index, donor in enumerate(donors):
    # donorObj = MetadataDocument(donor)
    dumpJsonToFile(json.dumps(donor), dir, "donor_"+str(index))
    ingest_api.createDonor(json.dumps(donor))

for index, assay in enumerate(assays):
    # assayObj = MetadataDocument(assay)
    dumpJsonToFile(json.dumps(assay), dir, "assay_"+str(index))
    ingest_api.createAssay(json.dumps(assay))

ingest_api.finishSubmission()

print "All done!"
wb.close()
exit()
